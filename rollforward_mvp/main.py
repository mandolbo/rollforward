"""
롤포워딩 MVP - 백데이터 시스템 버전
목표: 전기 조서 내 백데이터 자동 식별 및 색상 표시

이 프로그램이 하는 일:
1. 전기 조서를 읽어서 워크시트를 본 조서/백데이터로 분류
2. 백데이터 셀을 빨간색으로 표시하여 롤포워딩 대상 표시
3. 롤포워딩 후 성공/실패에 따라 초록색/빨간색으로 상태 표시
4. 미처리 셀에 대한 상세 리포트 제공

왜 이 프로그램이 필요한가?
- 전기 조서 내 백데이터를 자동으로 식별하기 위해
- 롤포워딩 대상을 시각적으로 명확히 하기 위해
- 수작업 조정이 필요한 항목을 체계적으로 관리하기 위해
"""

# 왜 import를 하는가?
# 다른 Python 파일에서 만든 함수들을 이 파일에서 사용하기 위해
# 마치 도구상자에서 필요한 도구를 꺼내는 것과 같음

from table_finder import find_tables      # Excel에서 테이블을 찾는 함수
from header_matcher import match_headers   # 헤더를 매칭하는 함수
from file_selector import (                # 파일 선택 UI 기능들
    select_previous_file,                   # 전기 조서 파일 선택
    select_current_folder,                  # 당기 PBC 폴더 선택
    get_excel_files_in_folder,             # 폴더에서 Excel 파일 찾기
    show_selection_summary,                 # 선택 사항 요약 표시
    confirm_selection,                      # 사용자 확인 받기
    select_main_worksheets                  # 워크시트 분류 선택 (본 조서/백데이터)
)
from file_updater import (
    update_file,                            # 파일 업데이트 (기존)
    mark_back_data_red,                     # 백데이터 셀 빨간색 표시
    update_rollforward_status,              # 롤포워딩 상태 업데이트
    generate_manual_adjustment_report       # 수동 조정 리포트 생성
)
from memory_efficient_copy import (
    worksheet_full_replace,                 # 워크시트 전체 교체 (프로세스 A)
    copy_worksheet_like_ctrl_cv             # Ctrl+C/V 방식 복사 (프로세스 A)
)
from rollforward_log import (
    create_rollforward_log_worksheet        # 롤포워딩 로그 워크시트 생성
)
import os
import re
import sys
import logging
from datetime import datetime
from pathlib import Path

# =================================================================
# 로그 관리 및 파일 선택 기능
# =================================================================

# LogCapture 클래스와 TeeOutput 클래스 삭제됨 - Excel 로그 워크시트로 대체

# =================================================================
# 프로세스 A: 백데이터 워크시트 처리 함수들
# =================================================================

def detect_backdata_worksheets(file_path):
    """
    독립 워크시트 자동 감지 - 본 조서 이외의 완전히 독립된 워크시트 식별
    
    신뢰도 점수 계산 (70% 이상만 반환):
    - 워크시트명 패턴 분석 (40% 가중치)
    - 내용 구조 분석 (35% 가중치) 
    - 데이터 밀도 분석 (25% 가중치)
    
    Args:
        file_path (str): 전기 조서 파일 경로
        
    Returns:
        list: [{"name": 워크시트명, "confidence": 신뢰도, "reason": 감지 근거}, ...]
    """
    
    if not os.path.exists(file_path):
        print(f"[detect_backdata_worksheets] 파일이 존재하지 않습니다: {file_path}")
        return []
    
    try:
        from openpyxl import load_workbook
        
        # 백데이터 워크시트 패턴들 (정규표현식)
        backdata_patterns = [
            r'별도.*',           # '별도BS', '별도손익' 등
            r'.*백.*데이터.*',   # '백데이터', '매출백데이터' 등
            r'.*분석.*',         # '매출분석', '비용분석', '분석표' 등
            r'.*명세.*',         # '자산명세', '부채명세' 등
            r'.*내역.*',         # '매출내역', '비용내역' 등
            r'.*상세.*',         # '상세내역', '상세분석' 등
            r'.*조정.*',         # '조정사항', '조정내역' 등
        ]
        
        # 본 조서 패턴들 (제외 대상)
        main_patterns = [
            r'^BS$', r'^PL$', r'^CF$',
            r'^재무상태표$', r'^손익계산서$', r'^현금흐름표$',
            r'^대차대조표$', r'^포괄손익계산서$'
        ]
        
        # 파일 접근 권한 체크
        if not os.access(file_path, os.R_OK):
            print(f"[detect_backdata_worksheets] 파일 읽기 권한 없음: {file_path}")
            return []
        
        wb = load_workbook(file_path, read_only=True, data_only=True)
        backdata_candidates = []
        
        for ws in wb.worksheets:
            ws_name = ws.title
            
            # 1단계: 본 조서인지 확인 (본 조서는 제외)
            is_main_sheet = any(re.match(pattern, ws_name, re.IGNORECASE) for pattern in main_patterns)
            if is_main_sheet:
                continue
                
            # 2단계: 백데이터 패턴 매칭 점수 계산 (40% 가중치)
            pattern_score = 0.0
            matched_patterns = []
            
            for pattern in backdata_patterns:
                if re.search(pattern, ws_name, re.IGNORECASE):
                    pattern_score = 0.4  # 패턴 매칭 시 40% 점수
                    matched_patterns.append(pattern)
                    break
            
            # 3단계: 내용 구조 분석 (35% 가중치)
            structure_score = analyze_worksheet_structure(ws)
            
            # 4단계: 데이터 밀도 분석 (25% 가중치)
            density_score = analyze_data_density(ws)
            
            # 총 신뢰도 계산
            total_confidence = pattern_score + (structure_score * 0.35) + (density_score * 0.25)
            
            # 70% 이상만 백데이터로 판정
            if total_confidence >= 0.7:
                backdata_candidates.append({
                    "name": ws_name,
                    "confidence": total_confidence,
                    "reason": f"패턴매칭: {matched_patterns}, 구조점수: {structure_score:.2f}, 밀도점수: {density_score:.2f}"
                })
                print(f"[detect_backdata_worksheets] 백데이터 감지: '{ws_name}' (신뢰도: {total_confidence:.1%})")
        
        wb.close()
        
        # 신뢰도 순으로 정렬
        backdata_candidates.sort(key=lambda x: x['confidence'], reverse=True)
        
        print(f"[detect_backdata_worksheets] 감지 완료: {len(backdata_candidates)}개 워크시트")
        return backdata_candidates
        
    except PermissionError as e:
        print(f"[detect_backdata_worksheets] 파일 접근 권한 오류: {file_path}")
        print(f"[detect_backdata_worksheets] 해결방법: Excel에서 파일을 닫고 다시 시도하세요")
        return []
    except Exception as e:
        print(f"[detect_backdata_worksheets] 감지 실패: {e}")
        return []

def analyze_worksheet_structure(ws):
    """워크시트 구조 분석하여 백데이터 가능성 점수 반환"""
    try:
        # 간단한 구조 분석: 테이블 형태인지, 복잡한 레이아웃인지
        dimension = ws.calculate_dimension()
        if not dimension:
            return 0.0
            
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(dimension)
        
        # 데이터 범위가 적절한지 확인 (너무 작으면 제목이나 메모일 수 있음)
        if max_row - min_row < 3 or max_col - min_col < 2:
            return 0.3  # 낮은 점수
            
        # 적절한 크기의 구조화된 데이터로 판정
        return 0.8
        
    except Exception:
        return 0.5  # 분석 실패 시 중간 점수

def analyze_data_density(ws):
    """워크시트 데이터 밀도 분석"""
    try:
        dimension = ws.calculate_dimension()
        if not dimension:
            return 0.0
            
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(dimension)
        
        total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        non_empty_cells = 0
        
        # 샘플링으로 밀도 확인 (성능 최적화)
        sample_size = min(100, total_cells)
        step = max(1, total_cells // sample_size)
        
        current_cell = 0
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, 
                               min_col=min_col, max_col=max_col, values_only=True):
            for cell_value in row:
                current_cell += 1
                if current_cell % step == 0:  # 샘플링
                    if cell_value is not None and str(cell_value).strip():
                        non_empty_cells += 1
                if current_cell >= sample_size:
                    break
            if current_cell >= sample_size:
                break
        
        density = non_empty_cells / sample_size if sample_size > 0 else 0
        
        # 밀도가 30-80% 사이면 백데이터에 적합한 구조
        if 0.3 <= density <= 0.8:
            return 0.9
        elif 0.1 <= density <= 0.9:
            return 0.6
        else:
            return 0.2
            
    except Exception:
        return 0.5

def find_matching_worksheet(worksheet_name, current_files):
    """
    당기 PBC 파일들에서 매칭되는 워크시트 찾기
    
    매칭 우선순위:
    1. 정확한 이름 매칭 (100%)
    2. 대소문자 무시 매칭 (95%)
    3. 공백 제거 매칭 (90%)
    4. 유사도 기반 매칭 (80% 이상)
    
    Args:
        worksheet_name (str): 찾을 워크시트 이름
        current_files (list): 당기 PBC 파일 리스트
        
    Returns:
        dict: {"file_path": 파일경로, "sheet_name": 시트명, "confidence": 신뢰도} 또는 None
    """
    
    try:
        from openpyxl import load_workbook
        
        best_match = None
        best_confidence = 0.0
        
        for file_path in current_files:
            try:
                wb = load_workbook(file_path, read_only=True)
                
                for ws in wb.worksheets:
                    current_sheet_name = ws.title
                    
                    # 1. 정확한 매칭
                    if worksheet_name == current_sheet_name:
                        wb.close()
                        return {
                            "file_path": file_path,
                            "sheet_name": current_sheet_name,
                            "confidence": 1.0
                        }
                    
                    # 2. 대소문자 무시 매칭
                    if worksheet_name.lower() == current_sheet_name.lower():
                        if 0.95 > best_confidence:
                            best_match = {
                                "file_path": file_path,
                                "sheet_name": current_sheet_name,
                                "confidence": 0.95
                            }
                            best_confidence = 0.95
                    
                    # 3. 공백 제거 매칭
                    clean_target = re.sub(r'\s+', '', worksheet_name)
                    clean_current = re.sub(r'\s+', '', current_sheet_name)
                    if clean_target.lower() == clean_current.lower():
                        if 0.9 > best_confidence:
                            best_match = {
                                "file_path": file_path,
                                "sheet_name": current_sheet_name,
                                "confidence": 0.9
                            }
                            best_confidence = 0.9
                
                wb.close()
                
            except Exception as e:
                print(f"[find_matching_worksheet] 경고: 파일 처리 오류 ({file_path}): {e}")
                continue
        
        return best_match if best_confidence >= 0.8 else None
        
    except Exception as e:
        print(f"[find_matching_worksheet] ❌ 매칭 실패: {e}")
        return None

def find_matching_worksheet_optimized(worksheet_name, current_files):
    """
    ⚡ 성능 최적화된 매칭 함수
    
    개선사항:
    1. 파일명 기반 1차 필터링 (성능 향상)
    2. 에러 처리 강화 (파일 잠금, 권한 문제)
    3. 한글 매칭 강화 (띄어쓰기, 특수문자)
    4. 매칭 과정 로깅 상세화
    
    Args:
        worksheet_name (str): 찾을 워크시트 이름
        current_files (list): 당기 PBC 파일 리스트
        
    Returns:
        dict: {"file_path": 파일경로, "sheet_name": 시트명, "confidence": 신뢰도} 또는 None
    """
    
    if not worksheet_name or not current_files:
        print(f"[find_matching_worksheet_optimized] ⚠️ 입력 매개변수 오류")
        return None
    
    try:
        from openpyxl import load_workbook
        import difflib
        
        print(f"[find_matching_worksheet_optimized] '{worksheet_name}' 매칭 시작...")
        
        best_match = None
        best_confidence = 0.0
        
        # 1단계: 파일명 기반 1차 필터링 (성능 최적화)
        prioritized_files = []
        other_files = []
        
        worksheet_clean = re.sub(r'[\s\-_]+', '', worksheet_name.lower())
        
        for file_path in current_files:
            filename = os.path.basename(file_path).lower()
            filename_clean = re.sub(r'[\s\-_\.xlsx]+', '', filename)
            
            # 파일맅에 워크시트명이 포함되어 있으면 우선 처리
            if worksheet_clean in filename_clean or filename_clean in worksheet_clean:
                prioritized_files.append(file_path)
                print(f"[find_matching_worksheet_optimized]    🎯 1차 우선 파일: {os.path.basename(file_path)}")
            else:
                other_files.append(file_path)
        
        # 우선 파일들을 먼저 처리
        files_to_process = prioritized_files + other_files
        
        for file_path in files_to_process:
            try:
                print(f"[find_matching_worksheet_optimized]    파일 처리 중: {os.path.basename(file_path)}")
                
                # 파일 존재 여부만 체크 (os.access는 클라우드 동기화에서 부정확할 수 있음)
                if not os.path.exists(file_path):
                    print(f"[find_matching_worksheet_optimized]    파일이 존재하지 않음: {file_path}")
                    continue
                
                # 파일 잠금 및 대기 시간 처리
                try:
                    wb = load_workbook(file_path, read_only=True)
                except PermissionError:
                    print(f"[find_matching_worksheet_optimized]    파일 잠김 (사용 중): {os.path.basename(file_path)}")
                    continue
                except Exception as file_error:
                    print(f"[find_matching_worksheet_optimized]    파일 로드 오류: {os.path.basename(file_path)} - {file_error}")
                    continue
                
                for ws in wb.worksheets:
                    current_sheet_name = ws.title
                    
                    # 1. 정확한 매칭 (100%)
                    if worksheet_name == current_sheet_name:
                        wb.close()
                        print(f"[find_matching_worksheet_optimized]    🎆 정확한 매칭 발견: {current_sheet_name}")
                        return {
                            "file_path": file_path,
                            "sheet_name": current_sheet_name,
                            "confidence": 1.0
                        }
                    
                    # 2. 대소문자 무시 매칭 (95%)
                    if worksheet_name.lower() == current_sheet_name.lower():
                        if 0.95 > best_confidence:
                            best_match = {
                                "file_path": file_path,
                                "sheet_name": current_sheet_name,
                                "confidence": 0.95
                            }
                            best_confidence = 0.95
                            print(f"[find_matching_worksheet_optimized]    🔤 대소문자 무시 매칭: {current_sheet_name} (95%)")
                    
                    # 3. 공백/특수문자 제거 매칭 (90%)
                    clean_target = re.sub(r'[\s\-_　]+', '', worksheet_name)
                    clean_current = re.sub(r'[\s\-_　]+', '', current_sheet_name)
                    if clean_target.lower() == clean_current.lower():
                        if 0.9 > best_confidence:
                            best_match = {
                                "file_path": file_path,
                                "sheet_name": current_sheet_name,
                                "confidence": 0.9
                            }
                            best_confidence = 0.9
                            print(f"[find_matching_worksheet_optimized]    🧽 공백 제거 매칭: {current_sheet_name} (90%)")
                    
                    # 4. 유사도 기반 매칭 (85% 이상)
                    if len(clean_target) > 2 and len(clean_current) > 2:
                        similarity = difflib.SequenceMatcher(None, clean_target.lower(), clean_current.lower()).ratio()
                        if similarity >= 0.85:
                            confidence_score = 0.8 + (similarity - 0.85) * 0.4  # 0.8-0.84 범위
                            if confidence_score > best_confidence:
                                best_match = {
                                    "file_path": file_path,
                                    "sheet_name": current_sheet_name,
                                    "confidence": confidence_score
                                }
                                best_confidence = confidence_score
                                print(f"[find_matching_worksheet_optimized]    📊 유사도 매칭: {current_sheet_name} ({confidence_score:.1%})")
                
                wb.close()
                
                # 우선 파일에서 좋은 매칭을 찾았으면 조기 종료
                if file_path in prioritized_files and best_confidence >= 0.9:
                    print(f"[find_matching_worksheet_optimized]    ⚡ 우선 파일에서 좋은 매칭 발견, 조기 종료")
                    break
                
            except Exception as e:
                print(f"[find_matching_worksheet_optimized]    파일 처리 오류 ({os.path.basename(file_path)}): {e}")
                continue
        
        # 결과 반환
        if best_match and best_confidence >= 0.8:
            print(f"[find_matching_worksheet_optimized] ✅ 최종 매칭 성공: '{worksheet_name}' → '{best_match['sheet_name']}' ({best_confidence:.1%})")
            return best_match
        else:
            print(f"[find_matching_worksheet_optimized] 매칭 실패: '{worksheet_name}' (최고 신뢰도: {best_confidence:.1%})")
            return None
        
    except Exception as e:
        print(f"[find_matching_worksheet_optimized] 매칭 처리 오류: {e}")
        return None

def copy_backdata_worksheets_corrected(backdata_worksheets, target_file, source_files):
    """
    ✅ 수정된 버전: 올바른 롤포워딩 방향 복사
    
    기존 문제: 전기 조서 → 당기 PBC (잘못된 방향)
    수정 후: 당기 PBC → 전기 조서 (올바른 롤포워딩)
    
    Args:
        backdata_worksheets (list): 백데이터 워크시트 정보 리스트
        target_file (str): 전기 조서 파일 경로 (타곃 - 복사될 곳)
        source_files (list): 당기 PBC 파일 리스트 (소스 - 복사할 데이터)
        
    Returns:
        dict: 처리 결과 {"success": [...], "failed": [...], "no_source": [...]}
    """
    
    results = {
        "success": [],
        "failed": [],
        "no_source": []  # source가 없음 (기존 no_target에서 수정)
    }
    
    print(f"[copy_backdata_worksheets_corrected] 🔄 올바른 롤포워딩 시작: {len(backdata_worksheets)}개")
    print(f"[copy_backdata_worksheets_corrected] 🎯 타곃: {target_file}")
    print(f"[copy_backdata_worksheets_corrected] 📁 소스: {len(source_files)}개 당기 PBC 파일")
    
    for worksheet_info in backdata_worksheets:
        worksheet_name = worksheet_info['name']
        confidence = worksheet_info['confidence']
        
        print(f"[copy_backdata_worksheets_corrected] 📋 처리 중: '{worksheet_name}' (신뢰도: {confidence:.1%})")
        
        # 1단계: 당기 PBC 파일에서 매칭되는 워크시트 찾기
        source_match = find_matching_worksheet_optimized(worksheet_name, source_files)
        
        if not source_match:
            print(f"[copy_backdata_worksheets_corrected] ⚠️ 경고: 당기 PBC에서 매칭되는 워크시트를 찾을 수 없음: {worksheet_name}")
            results["no_source"].append(worksheet_name)
            continue
        
        source_file = source_match['file_path']
        source_sheet = source_match['sheet_name']
        match_confidence = source_match['confidence']
        
        print(f"[copy_backdata_worksheets_corrected]    🎯 매칭 발견: {source_sheet} (PBC: {os.path.basename(source_file)}, 신뢰도: {match_confidence:.1%})")
        
        # 2단계: 올바른 방향 복사 실행 (당기 PBC → 전기 조서)
        try:
            success = worksheet_full_replace(
                source_file=source_file,      # 당기 PBC 파일 (소스)
                source_sheet=source_sheet,    # 당기 PBC 워크시트 (소스)
                target_file=target_file,      # 전기 조서 파일 (타곃)
                target_sheet=worksheet_name,  # 전기 조서 워크시트 (타곃)
                preserve_formulas=False       # 값 복사 (수식 보존 안함)
            )
            
            if success:
                print(f"[copy_backdata_worksheets_corrected]    ✅ 복사 성공: {source_sheet} ({os.path.basename(source_file)}) → {worksheet_name} (전기조서)")
                results["success"].append({
                    "source": source_sheet,
                    "source_file": source_file,
                    "target": worksheet_name,
                    "target_file": target_file,
                    "confidence": match_confidence
                })
            else:
                print(f"[copy_backdata_worksheets_corrected]    ❌ 복사 실패: {worksheet_name}")
                results["failed"].append(worksheet_name)
                
        except Exception as e:
            print(f"[copy_backdata_worksheets_corrected]    ❌ 복사 중 오류: {worksheet_name} - {e}")
            results["failed"].append(worksheet_name)
    
    # 결과 요약
    success_count = len(results["success"])
    failed_count = len(results["failed"])
    no_source_count = len(results["no_source"])
    
    print(f"[copy_backdata_worksheets_corrected] 📋 처리 결과:")
    print(f"[copy_backdata_worksheets_corrected]    ✅ 성공: {success_count}개")
    print(f"[copy_backdata_worksheets_corrected]    ❌ 실패: {failed_count}개")
    print(f"[copy_backdata_worksheets_corrected]    ⚠️ 소스 없음: {no_source_count}개")
    
    return results

def copy_backdata_worksheets(backdata_worksheets, previous_file, current_files):
    """
    memory_efficient_copy.py 연동하여 워크시트 전체를 통째로 복사
    
    Args:
        backdata_worksheets (list): 백데이터 워크시트 정보 리스트
        previous_file (str): 전기 조서 파일 경로
        current_files (list): 당기 PBC 파일 리스트
        
    Returns:
        dict: 처리 결과 {"success": [...], "failed": [...], "no_target": [...]}
    """
    
    results = {
        "success": [],
        "failed": [],
        "no_target": []
    }
    
    print(f"[copy_backdata_worksheets] 워크시트 복사 시작: {len(backdata_worksheets)}개")
    
    for worksheet_info in backdata_worksheets:
        worksheet_name = worksheet_info['name']
        confidence = worksheet_info['confidence']
        
        print(f"[copy_backdata_worksheets] 📋 처리 중: '{worksheet_name}' (신뢰도: {confidence:.1%})")
        
        # 1단계: 매칭되는 워크시트 찾기
        target_match = find_matching_worksheet(worksheet_name, current_files)
        
        if not target_match:
            print(f"[copy_backdata_worksheets] 경고: 매칭되는 대상 워크시트를 찾을 수 없음: {worksheet_name}")
            results["no_target"].append(worksheet_name)
            continue
        
        target_file = target_match['file_path']
        target_sheet = target_match['sheet_name']
        match_confidence = target_match['confidence']
        
        print(f"[copy_backdata_worksheets]    🎯 매칭 발견: {target_sheet} (신뢰도: {match_confidence:.1%})")
        
        # 2단계: 워크시트 복사 실행
        try:
            success = worksheet_full_replace(
                source_file=previous_file,
                source_sheet=worksheet_name,
                target_file=target_file,
                target_sheet=target_sheet,
                preserve_formulas=False  # 값 복사 (수식 보존 안함)
            )
            
            if success:
                print(f"[copy_backdata_worksheets]    ✅ 복사 성공: {worksheet_name} → {target_sheet}")
                results["success"].append({
                    "source": worksheet_name,
                    "target": target_sheet,
                    "file": target_file,
                    "confidence": match_confidence
                })
            else:
                print(f"[copy_backdata_worksheets]    ❌ 복사 실패: {worksheet_name}")
                results["failed"].append(worksheet_name)
                
        except Exception as e:
            print(f"[copy_backdata_worksheets]    ❌ 복사 중 오류: {worksheet_name} - {e}")
            results["failed"].append(worksheet_name)
    
    return results

def display_automatic_processing_info(backdata_worksheets):
    """
    자동 분류 시스템 정보 표시
    
    Args:
        backdata_worksheets (list): 감지된 백데이터 워크시트 정보
    """
    
    if not backdata_worksheets:
        return
    
    print("\n" + "="*60)
    print("🎯 자동 분류 시스템")
    print("="*60)
    
    print(f"\n📋 감지된 백데이터 워크시트: {len(backdata_worksheets)}개")
    for i, ws_info in enumerate(backdata_worksheets, 1):
        print(f"   {i}. '{ws_info['name']}' (신뢰도: {ws_info['confidence']:.1%}) → 워크시트 전체 복사")
    
    print(f"\n💡 처리 방식:")
    print("   📄 백데이터 워크시트 → 프로세스 A (워크시트 전체 복사)")
    print("   📊 일반 워크시트 → 프로세스 B (테이블 단위 처리)")
    print("   🚫 이중 처리 방지 → 복사된 워크시트는 테이블 처리에서 자동 제외")

def print_processing_summary_corrected(process_a_results=None, process_b_results=None):
    """
    ✅ 수정된 버전: 올바른 복사 방향이 반영된 결과 출력
    
    Args:
        process_a_results (dict): 프로세스 A 결과
        process_b_results (dict): 프로세스 B 결과 (선택적)
    """
    
    print("\n" + "="*80)
    print("📊 수정된 백데이터 처리 결과 종합 (올바른 롤포워딩)")
    print("="*80)
    
    if process_a_results:
        success_count = len(process_a_results.get('success', []))
        failed_count = len(process_a_results.get('failed', []))
        no_source_count = len(process_a_results.get('no_source', []))  # no_target에서 변경
        
        print(f"\n🔄 프로세스 A (올바른 롤포워딩: 당기 PBC → 전기 조서):")
        print(f"   ✅ 성공: {success_count}개")
        if success_count > 0:
            for result in process_a_results['success']:
                source_file_name = os.path.basename(result.get('source_file', ''))
                print(f"      - 📁 {result['source']} ({source_file_name}) → 📊 {result['target']} (전기조서)")
        
        print(f"   ❌ 실패: {failed_count}개")
        if failed_count > 0:
            for failed in process_a_results['failed']:
                print(f"      - ⚠️ {failed} (복사 오류 발생)")
        
        print(f"   🔍 소스 없음: {no_source_count}개")
        if no_source_count > 0:
            for no_source in process_a_results['no_source']:
                print(f"      - 📄 {no_source} (당기 PBC에서 매칭되는 워크시트 없음)")
        
        # 성공률 계산
        total_count = success_count + failed_count + no_source_count
        success_rate = (success_count / total_count * 100) if total_count > 0 else 0
        print(f"\n   📊 성공률: {success_count}/{total_count} ({success_rate:.1f}%)")
        
        if success_rate >= 80:
            print(f"   🎆 우수! 대부분의 백데이터가 성공적으로 롤포워드되었습니다.")
        elif success_rate >= 50:
            print(f"   🔧 보통: 일부 수동 조정이 필요합니다.")
        else:
            print(f"   ⚠️ 주의: 대부분의 백데이터에 대해 수동 처리가 필요합니다.")
    
    if process_b_results:
        print(f"\n📊 프로세스 B (기존 테이블 단위 처리):")
        print(f"   📊 기존 테이블 처리 시스템이 실행되었습니다")
        print(f"   💡 상세 결과는 위의 처리 로그를 참조하세요")
    
    print("\n💡 다음 조치 사항:")
    if process_a_results:
        if process_a_results.get('failed'):
            print("   ❗ 실패한 워크시트는 프로세스 B (테이블 단위)로 자동 처리됩니다")
        if process_a_results.get('no_source'):
            print("   📄 매칭되지 않은 워크시트는 당기 PBC 파일명과 워크시트명을 확인해주세요")
    
    print("   📊 Excel 파일을 열어서 롤포워딩 결과를 확인해보세요")
    print("   ✅ 색상 의미: 🟢 초록색 = 처리 완료, 🔴 빨간색 = 수동 조정 필요")
    print("   📝 로그 파일이 당기 PBC 폴더에 저장되었습니다")
    
def print_processing_summary(process_a_results=None, process_b_results=None):
    """기존 함수 호환성을 위한 래퍼"""
    return print_processing_summary_corrected(process_a_results, process_b_results)

def main():
    """
    백데이터 시스템 메인 함수: 두 독립적인 백데이터 처리 프로세스 통합
    
    새로운 통합 워크플로우 (v3.0):
    1. 전기 조서 파일 선택
    2. 프로세스 A: 독립 워크시트 자동 감지 및 전체 복사 (신규)
    3. 프로세스 B: 워크시트 분류 및 테이블 단위 처리 (기존)
    4. 두 프로세스 결과 통합 및 상태 표시
    5. 성공/실패에 따른 색상 피드백 시스템
    6. 미처리 항목에 대한 상세 리포트 생성
    7. 로그 파일 자동 생성 및 저장
    
    왜 이 통합 시스템이 혁신적인가?
    프로세스 A: 본 조서 이외의 독립된 워크시트를 통째로 복사
    프로세스 B: 본 조서 내부의 테이블을 선별적으로 처리
    하이브리드: 두 방식을 조합하여 완전한 백데이터 처리 실현
    """
    
    # 이전 txt 로그 기능 제거됨 - Excel 로그 워크시트 사용
    
    try:
        print("롤포워딩 MVP - 백데이터 시스템 시작!")  # 사용자에게 프로그램이 시작되었다고 알려줌
        print("전기 조서 파일과 워크시트를 선택해주세요.\n")
        
        # =================================================================
        # 새로운 기능: 사용자가 직접 파일과 폴더를 선택
        # =================================================================
        # 왜 하드코딩에서 사용자 선택으로 바꾸는가?
        # - 다양한 환경에서 프로그램을 사용할 수 있게 하기 위해
        # - 실제 업무에서는 파일 위치가 매번 다르기 때문에
        # - 사용자 친화적인 프로그램을 만들기 위해
        
        # 1단계: 전기 조서 파일 선택
        print("1. 전기 조서 파일을 선택해주세요...")
        previous_file = select_previous_file()
        
        # 사용자가 파일 선택을 취소한 경우
        if not previous_file:
            print("전기 조서 파일 선택이 취소되었습니다.")
            print("프로그램을 종료합니다.")
            return  # main 함수 종료 (프로그램 끝)
        
        # =================================================================
        # 1.5단계: 혁신적 워크시트 분류 시스템 (백데이터 자동 식별의 핵심)
        # =================================================================
        # 왜 이 단계가 혁신적인가?
        # 기존: 전기 조서 → 당기 PBC 단순 복사 (어느 데이터가 백데이터인지 모름)
        # 신규: 전기 조서 내에서 본 조서 vs 백데이터를 사용자가 직접 분류
        #       → 백데이터만 선별적으로 롤포워딩 대상으로 표시 및 처리
        
        # =================================================================
        # 1.5단계: 워크시트 분류 (사용자 기대 UI 순서대로)
        # =================================================================
        
        print("1.5 워크시트를 본 조서와 백데이터로 분류해주세요...")
        print("     💡 본 조서: 최종 재무제표 (손익계산서, 재무상태표 등)")
        print("     🔴 백데이터: 상세 분석 자료 (매출분석, 비용분석, 조정사항 등)")
        
        # 프로세스 B: 워크시트 분류 (user_confirmation.py UI 사용)
        main_worksheets, back_data_worksheets = select_main_worksheets(previous_file)
        
        print(f"[main.main] 📋 본 조서 워크시트: {len(main_worksheets)}개")
        print(f"[main.main] 🔴 백데이터 워크시트: {len(back_data_worksheets)}개")
        
        # 2단계: 당기 PBC 폴더 선택 
        print("\n2. 당기 PBC 폴더를 선택해주세요...")
        current_folder = select_current_folder()
        
        if not current_folder:
            print("❌ 당기 PBC 폴더 선택이 취소되었습니다.")
            print("프로그램을 종료합니다.")
            return
        
        # 이전 txt 로그 캡처 기능 제거됨 - Excel 로그 워크시트로 대체
        
        # 당기 파일들 목록 가져오기
        current_files = get_excel_files_in_folder(current_folder)
        
        if not current_files:
            print("❌ 선택한 폴더에 Excel 파일이 없습니다.")
            print("다른 폴더를 선택해주세요.")
            return
        
        # 3단계: 선택 사항 요약 및 Roll-Forwarding 확인
        show_selection_summary(previous_file, current_folder, current_files)
        
        # 사용자가 계속 진행하지 않겠다고 한 경우
        if not confirm_selection():
            print("❌ 사용자가 작업을 취소했습니다.")
            print("프로그램을 종료합니다.")
            return
        
        # =================================================================
        # 추가: 프로세스 A - 독립 워크시트 자동 감지 및 처리 (백데이터 중에서)
        # =================================================================
        
        process_a_results = None
        
        if back_data_worksheets:
            print("\n🔍 프로세스 A: 백데이터 워크시트 전체 복사 처리 시작...")
            print(f"[main.main] 📋 처리 대상 백데이터 워크시트: {len(back_data_worksheets)}개")
            
            # ✅ 수정: 모든 백데이터 워크시트를 프로세스 A 대상으로 처리
            # 기존 문제: 백데이터 중에서 또 다시 패턴 매칭하는 중복 로직
            # 수정 후: 사용자가 분류한 모든 백데이터 워크시트를 처리 대상으로 인식
            all_backdata_worksheets = []
            for ws_name in back_data_worksheets:
                all_backdata_worksheets.append({
                    "name": ws_name,
                    "confidence": 1.0,  # 사용자가 직접 분류했으므로 100% 신뢰도
                    "reason": "사용자 직접 분류된 백데이터 워크시트"
                })
                print(f"[main.main]    - '{ws_name}' (사용자 분류 확정)")
            
            print(f"[main.main] 🎯 총 처리 대상: {len(all_backdata_worksheets)}개 워크시트")
            
            # 자동 분류 시스템 정보 표시
            display_automatic_processing_info(all_backdata_worksheets)
            
            # 자동 분류: 백데이터 워크시트는 전체 복사 (프로세스 A)
            print("\n[main.main] 프로세스 A 실행: 백데이터 워크시트 자동 전체 복사")
            print("[main.main] 💡 당기 PBC 파일 → 전기 조서 워크시트로 롤포워딩 시작")
            
            process_a_results = copy_backdata_worksheets_corrected(
                all_backdata_worksheets, previous_file, current_files
            )
            
            # 프로세스 A에서 성공한 워크시트 목록 추출
            successfully_copied_worksheets = []
            if process_a_results and 'success' in process_a_results:
                for result in process_a_results['success']:
                    successfully_copied_worksheets.append(result['target'])  # 워크시트 이름
            
            print(f"[main.main] 🎯 프로세스 A 완료: {len(successfully_copied_worksheets)}개 워크시트 복사 성공")
            if successfully_copied_worksheets:
                print(f"[main.main]    복사된 워크시트: {', '.join(successfully_copied_worksheets)}")
        else:
            print("\n💡 백데이터 워크시트가 감지되지 않았습니다.")
            print("[main.main] 프로세스 A 건너뛰기 - 일반 워크시트 테이블 단위 처리만 실행")
            successfully_copied_worksheets = []
        
        print("[main.main] \n✅ 파일 선택 완료! 롤포워딩을 시작합니다...\n")
        
        # =================================================================
        # 1단계: 전기 조서(작년 데이터)에서 테이블 찾기
        # =================================================================
        # 왜 이 단계가 필요한가?
        # Excel 파일은 여러 시트가 있고, 각 시트에는 제목, 공백, 실제 데이터가 섞여있음
        # 실제 데이터가 어디에 있는지 찾아야 복사할 수 있음
        
        print("[main.main] 📊 1단계: 테이블 찾는 중...")
        # 전체 테이블 찾기
        all_previous_tables = find_tables(previous_file)  # table_finder.py의 함수 호출
        
        # 프로세스 A에서 성공한 워크시트가 있는 경우, 해당 워크시트의 테이블 제외
        excluded_worksheets = []
        if 'successfully_copied_worksheets' in locals() and successfully_copied_worksheets:
            excluded_worksheets = successfully_copied_worksheets
            print(f"[main.main] 🎯 이미 복사된 워크시트 제외: {', '.join(excluded_worksheets)}")
        
        # 이미 복사된 워크시트의 테이블 필터링
        previous_tables = []
        for table in all_previous_tables:
            if table['sheet'] not in excluded_worksheets:
                previous_tables.append(table)
        
        print(f"[main.main]    → 전체 테이블 수: {len(all_previous_tables)}개")
        if excluded_worksheets:
            print(f"[main.main]    → 제외된 테이블 수: {len(all_previous_tables) - len(previous_tables)}개")
        print(f"[main.main]    → 프로세스 B 처리 대상: {len(previous_tables)}개")
        
        # =================================================================
        # 1.5단계: 백데이터 시각화 시스템 (롤포워딩 대상 명확화)
        # =================================================================
        # 혁신적 접근법:
        # 1) 백데이터 워크시트의 모든 테이블 셀을 빨간색으로 표시
        # 2) "롤포워딩 대상" 주석을 각 셀에 추가하여 용도 명시
        # 3) 나중에 처리 상태에 따라 초록색(성공) 또는 빨간색 유지(실패)
        # 
        # 사용자 경험 개선점:
        # - 어떤 데이터가 롤포워딩 대상인지 한 눈에 파악 가능
        # - Excel에서 직접 확인 가능한 시각적 피드백 제공
        # - 처리 결과를 색상으로 즉시 확인 가능
        
        red_cells_info = {}  # 빨간색으로 표시된 셀 정보 저장용 딕셔너리
        
        # 백데이터 시각화: 프로세스 B 대상 테이블만 (이미 복사된 워크시트 제외)
        remaining_back_data_worksheets = []
        if back_data_worksheets:
            for ws in back_data_worksheets:
                if ws not in excluded_worksheets:
                    remaining_back_data_worksheets.append(ws)
                    
        if remaining_back_data_worksheets:
            print("[main.main] 🔴 1.5단계: 백데이터 시각화 시작...")
            print("[main.main]    💡 프로세스 B 대상 테이블의 모든 셀을 빨간색으로 표시합니다")
            print("[main.main]    💡 이후 롤포워딩 성공시 초록색으로, 실패시 빨간색 유지됩니다")
            
            red_cells_info = mark_back_data_red(previous_file, remaining_back_data_worksheets, previous_tables)
            total_marked_cells = sum(len(cells) for cells in red_cells_info.values())
            print(f"[main.main]    ✅ 롤포워딩 대상으로 표시된 셀 수: {total_marked_cells:,}개")
            
            if total_marked_cells == 0:
                print("[main.main]    경고: 남은 백데이터 워크시트에서 테이블을 찾을 수 없습니다.")
                print("[main.main]    💡 AutoFilter를 적용하거나 명확한 헤더가 있는지 확인해주세요.")
        elif back_data_worksheets:
            print("[main.main] 🎯 1.5단계: 모든 백데이터 워크시트가 프로세스 A에서 처리됨")
            print("[main.main]    💡 프로세스 B에서 처리할 백데이터 워크시트가 없습니다")
            red_cells_info = {}
        
        # =================================================================
        # 2단계: 당기 파일들(올해 작업할 파일들) 하나씩 처리
        # =================================================================
        # 왜 하나씩 처리하는가?
        # 각 파일마다 구조가 다를 수 있고, 매칭되는 헤더도 다를 수 있기 때문
        
        print("[main.main] 2단계: 당기 파일들 처리 중...")
        # current_files는 이미 위에서 get_excel_files_in_folder()로 가져옴
        print(f"[main.main]    → 처리할 파일 수: {len(current_files)}개")
        
        # 왜 for 반복문을 사용하는가?
        # 여러 개의 파일을 같은 방식으로 처리하기 위해
        # 코드 중복을 피하고 효율적으로 작업하기 위해
        for current_file in current_files:
            print(f"[main.main]    🔍 처리 중: {current_file}")
            
            # =================================================================
            # 3단계: 헤더 매칭 (어느 컬럼이 어느 컬럼과 연결되는지 찾기)
            # =================================================================
            # 왜 헤더 매칭이 필요한가?
            # 전기 조서의 "매출액" 컬럼이 당기 파일의 어느 컬럼에 해당하는지 알아야
            # 정확한 위치에 데이터를 복사할 수 있기 때문
            
            # current_folder + current_file은 문자열을 합치는 것
            # 예: "test_files/current_folder/" + "workpaper1.xlsx" = "test_files/current_folder/workpaper1.xlsx"
            matches = match_headers(previous_tables, current_folder + current_file)
            
            # =================================================================
            # 4단계: 실제 파일 업데이트 (데이터 복사)
            # =================================================================
            # 왜 if문을 사용하는가?
            # 매칭이 성공했을 때만 업데이트를 진행하기 위해
            # 매칭이 실패했는데 업데이트하면 잘못된 데이터가 들어갈 수 있음
            
            if matches:  # matches가 있으면(매칭 성공했으면)
                # 올바른 데이터 흐름: Current PBC → 백데이터 sheets
                # current_pbc_path: 당기 PBC 파일 (소스)
                # previous_file: 전기 조서 파일 (백데이터 시트가 있는 대상)
                current_pbc_path = current_folder + current_file
                success = update_file(matches, current_pbc_path, previous_file)
                
                # 왜 삼항 연산자를 사용하는가?
                # success가 True면 "성공", False면 "실패"를 간단하게 표현하기 위해
                # 긴 if-else 구문보다 간결함
                print(f"[main.main]    ✅ 업데이트: {'성공' if success else '실패'}")
            else:  # matches가 없으면(매칭 실패했으면)
                print(f"[main.main]    경고: 매칭 실패: {current_file}")
        
        # =================================================================
        # 5단계: 지능형 상태 추적 및 리포트 시스템 (백데이터 시스템의 핵심)
        # =================================================================
        # 이 단계에서 무엇이 일어나는가?
        # 1) 처리 성공한 백데이터 셀 → 빨간색에서 초록색으로 변경
        # 2) 처리 실패한 백데이터 셀 → 빨간색 유지 (시각적 경고)
        # 3) 실패 원인별 상세 분석 리포트 생성
        # 4) 수동 조정 가이드 제공
        #
        # 사용자에게 제공하는 가치:
        # - 투명성: 모든 처리 상태를 한눈에 확인 가능
        # - 효율성: 실패한 항목만 집중적으로 수동 처리 가능  
        # - 추적성: 어떤 셀이 왜 실패했는지 상세 정보 제공
        
        if back_data_worksheets and red_cells_info:
            print("[main.main] 🟢 5단계: 백데이터 처리 상태 업데이트 시작...")
            print("[main.main]    💡 성공한 셀 → 초록색, 실패한 셀 → 빨간색 유지")
            
            # 성공적으로 매칭된 항목들 수집 (모든 당기 파일에서)
            successful_matches = []
            failed_table_matches = []  # 테이블 매칭 실패 정보 수집
            
            for current_file in current_files:
                matches = match_headers(previous_tables, current_folder + current_file)
                if matches:
                    successful_matches.extend(matches)
                    print(f"[main.main]    ✅ 테이블 매칭 성공: {os.path.basename(current_file)} ({len(matches)}개)")
                else:
                    failed_table_matches.append(current_file)
                    print(f"[main.main]    ❌ 테이블 매칭 실패: {os.path.basename(current_file)}")
            
            print(f"[main.main]    📊 총 성공한 헤더 매칭: {len(successful_matches)}개")
            
            # 상태 업데이트: 성공 → 초록색, 실패 → 빨간색 유지
            update_rollforward_status(previous_file, red_cells_info, successful_matches)
            
            # 지능형 수동 조정 리포트 생성
            print("[main.main] 📋 수동 조정 리포트 생성 중...")
            print("[main.main]    💡 실패 원인 분석 및 해결 가이드 제공")
            
            report = generate_manual_adjustment_report(previous_file, red_cells_info, successful_matches)
            
            if report:
                print("\n" + "="*70)
                print("📋 백데이터 처리 결과 및 수동 조정 가이드")
                print("="*70)
                print(report)
                print("="*70)
                print("💡 위 리포트를 참고하여 빨간색으로 남은 셀들을 수동으로 조정해주세요.")
            else:
                print("[main.main] 🎉 축하합니다! 모든 백데이터가 성공적으로 처리되었습니다!")
                print("[main.main] 💚 전기 조서의 모든 백데이터 셀이 초록색으로 표시되었습니다.")
        
# =================================================================
        # 최종 단계: 두 프로세스 통합 결과 출력
        # =================================================================
        
        print_processing_summary(process_a_results, {"status": "completed"})
        
        # =================================================================
        # 최종 단계: 롤포워딩 로그 워크시트 생성
        # =================================================================
        print("\n[main.main] 📊 6단계: 롤포워딩 로그 워크시트 생성 시작...")
        
        try:
            # 로그 데이터 수집
            from datetime import datetime
            from pathlib import Path
            
            log_data = {
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'target_file_name': Path(previous_file).name,
                'source_folder': current_folder,
                'total_worksheets': 0,
                'success_worksheets': 0,
                'failed_worksheets': 0,
                'total_tables': 0,  # 테이블 관련 필드 추가
                'success_tables': 0,  # 테이블 관련 필드 추가
                'failed_tables': 0,  # 테이블 관련 필드 추가
                'backdata_results': [],
                'table_results': [],
                'failures': [],
                'statistics': {
                    'total_time': 'N/A',
                    'copied_cells': 0,
                    'backup_files': 1,  # 영구 백업 파일
                    'processed_tables': len(previous_tables) if 'previous_tables' in locals() else 0
                }
            }
            
            # 프로세스 A 결과를 로그 데이터에 추가
            if process_a_results and 'success' in process_a_results:
                log_data['success_worksheets'] += len(process_a_results['success'])
                log_data['total_worksheets'] += len(process_a_results['success'])
                
                for success_item in process_a_results['success']:
                    log_data['backdata_results'].append({
                        'target_worksheet': success_item.get('target', 'N/A'),
                        'source_file': success_item.get('source', 'N/A'),
                        'source_worksheet': success_item.get('target', 'N/A'),  # 동일한 이름
                        'confidence': 1.0,  # 워크시트 전체 복사는 100% 신뢰도
                        'success': True
                    })
            
            if process_a_results and 'failed' in process_a_results:
                log_data['failed_worksheets'] += len(process_a_results['failed'])
                log_data['total_worksheets'] += len(process_a_results['failed'])
                
                for failed_item in process_a_results['failed']:
                    log_data['failures'].append({
                        'reason': f"백데이터 워크시트 '{failed_item}' 복사 실패",
                        'solution': "1. 소스 파일에서 해당 워크시트 존재 확인\n2. 파일 권한 및 Excel 프로그램 종료 확인"
                    })
            
            # 프로세스 B (테이블 단위 처리) 결과를 로그 데이터에 추가
            if 'successful_matches' in locals() and successful_matches:
                log_data['success_tables'] += len(successful_matches)
                log_data['total_tables'] += len(successful_matches)
                
                for match in successful_matches:
                    log_data['table_results'].append({
                        'target_worksheet': match.get('from_table', {}).get('sheet', 'N/A'),
                        'target_table_range': f"Row {match.get('from_table', {}).get('start_row', 'N/A')}",
                        'source_file': os.path.basename(current_folder + match.get('to_table', {}).get('file', 'N/A')),
                        'source_worksheet': match.get('to_table', {}).get('sheet', 'N/A'),
                        'source_table_range': f"Row {match.get('to_table', {}).get('start_row', 'N/A')}",
                        'matched_headers': f"{match.get('from_header', 'N/A')} ↔ {match.get('to_header', 'N/A')}",
                        'confidence': match.get('confidence', 1.0),
                        'success': True
                    })
            
            if 'failed_table_matches' in locals() and failed_table_matches:
                log_data['failed_tables'] += len(failed_table_matches)
                log_data['total_tables'] += len(failed_table_matches)
                
                for failed_file in failed_table_matches:
                    log_data['failures'].append({
                        'reason': f"테이블 헤더 매칭 실패: '{os.path.basename(failed_file)}'",
                        'solution': "1. 당기 PBC 파일과 전기 조서의 헤더명 일치 여부 확인\n2. 테이블에 명확한 헤더 행이 있는지 확인\n3. AutoFilter 적용 여부 확인"
                    })
            
            # 로그 워크시트 생성
            log_success = create_rollforward_log_worksheet(previous_file, log_data)
            
            if log_success:
                print("[main.main] ✅ 롤포워딩 로그 워크시트 생성 완료")
                print(f"[main.main]    💡 '{Path(previous_file).name}'에서 'Roll-Forward_Log_*' 워크시트를 확인하세요")
            else:
                print("[main.main] ⚠️ 롤포워딩 로그 워크시트 생성 실패")
                
        except Exception as log_error:
            print(f"[main.main] ⚠️ 로그 워크시트 생성 중 오류: {log_error}")
        
        print("\n[main.main] 🎉 두 독립적인 백데이터 처리 프로세스 통합 시스템 완료!")
        print("[main.main] 프로세스 A: 독립 워크시트 전체 복사")
        print("[main.main] 프로세스 B: 테이블 단위 선별적 처리")
        print("[main.main] 📊 Excel 파일을 열어서 결과를 확인해보세요:")
        print("[main.main]    🔴 빨간색 = 수동 조정 필요")  
        print("[main.main]    🟢 초록색 = 처리 완료")
        print("[main.main]    📋 로그 워크시트 = 전체 처리 내역 확인")
        print("[main.main] 💡 MVP 테스트 완료 - Phase 1 통합 구현 성공!")
        
    # 왜 except Exception을 사용하는가?
    # 어떤 종류의 오류든 잡아서 처리하기 위해
    # 사용자에게 친절한 에러 메시지와 해결 방법을 제시하기 위해
    except Exception as e:
        print(f"[main.main] ❌ 오류 발생: {e}")  # 구체적인 오류 내용을 보여줌
        print("[main.main] debug_collector.py 실행해서 ChatGPT에 문의하세요!")  # 해결 방법 안내
    
    finally:
        # 이전 txt 로그 캡처 종료 기능 제거됨
        pass

# 왜 이 if문이 필요한가?
# 이 파일이 직접 실행될 때만 main()을 호출하기 위해
# 다른 파일에서 이 파일을 import할 때는 main()이 자동으로 실행되지 않게 하기 위해
# Python의 관례적인 패턴임
if __name__ == "__main__":
    main()  # 메인 함수 실행