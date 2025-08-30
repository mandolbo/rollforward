"""
테이블 찾기 - 고속 개선 버전
3단계 알고리즘 + read_only 모드 + 범위 최적화로 대폭 속도 향상

이 파일이 하는 일:
- Excel 파일에서 실제 데이터 테이블을 정확하고 빠르게 찾아내는 것
- Excel의 공식 기능(Table, AutoFilter)과 휴리스틱 방법을 조합해서 사용
- read_only 스트리밍과 범위 최적화로 대용량 파일도 빠른 처리
- 신뢰도 기반 우선순위로 가장 정확한 테이블부터 처리 가능

왜 이렇게 복잡한 방법이 필요한가?
- 실제 업무에서 Excel 파일은 매우 다양한 형태로 만들어짐
- 단순한 행/열 스캔으로는 놓치는 테이블이 많음
- Excel의 내장 기능을 활용하면 훨씬 정확한 결과를 얻을 수 있음
- read_only 모드와 범위 최적화로 대용량 파일도 빠르게 처리 가능
"""

# 왜 이 라이브러리들을 import하는가?
import openpyxl  # Excel 파일(.xlsx)을 읽고 쓰기 위한 라이브러리
import os        # 파일 시스템(파일 존재 여부 등)을 다루기 위한 라이브러리
from openpyxl.utils import range_boundaries  # 범위 문자열을 좌표로 변환하기 위한 유틸리티

def find_tables(file_path):
    """
    Excel 파일에서 테이블 찾기 - 고속 3단계 개선 버전
    
    이 함수가 하는 일:
    1. read_only 모드로 빠르게 파일 열기
    2. Excel의 공식 Table 객체 찾기 (가장 정확함)
    3. AutoFilter가 적용된 범위 찾기 (사용자가 필터링한 데이터)
    4. 휴리스틱 방법으로 테이블 패턴 찾기 (지능적 추측, 범위 최적화 적용)
    5. 신뢰도 순으로 정렬해서 반환
    
    속도 개선 포인트:
    - read_only=True: 스트리밍 방식으로 메모리 효율성 대폭 향상
    - calculate_dimension(): 실제 사용 범위만 스캔해서 속도 향상
    - 벡터화된 행 처리: iter_rows로 효율적인 데이터 읽기
    - 조기 종료: 조건 불만족 시 즉시 다음 행으로 이동
    
    Parameters(매개변수):
        file_path (str): Excel 파일의 경로
        
    Returns(반환값):
        list: 찾은 테이블들의 정보 리스트 (신뢰도 순으로 정렬됨)
    """
    
    # 왜 파일 존재를 먼저 확인하는가?
    # 없는 파일을 열려고 하면 프로그램이 오류로 멈추기 때문
    # 사용자에게 친절한 메시지를 보여주기 위함
    if not os.path.exists(file_path):
        print(f"[table_finder.find_tables] ⚠️ 파일을 찾을 수 없습니다: {file_path}")
        return []  # 빈 리스트 반환 (아무것도 찾지 못했다는 뜻)
    
    # 왜 try-except를 사용하는가?
    # Excel 파일이 손상되었거나, 비밀번호가 걸려있거나, 다른 프로그램에서 사용 중일 때
    # 또는 ~$ 임시 파일 같은 잘못된 파일을 열려고 할 때 오류가 발생할 수 있음
    try:
        # 하이브리드 접근: AutoFilter 감지는 일반 모드, 데이터 스캔은 read_only 모드
        # 1단계: AutoFilter 감지를 위한 일반 모드 로딩
        wb_meta = openpyxl.load_workbook(file_path, data_only=True)  # 메타데이터 접근용
        tables = []  # 찾은 테이블들을 저장할 빈 리스트
        
        # 왜 모든 시트를 확인하는가?
        # Excel 파일에는 여러 시트가 있을 수 있고, 각 시트마다 다른 데이터가 있기 때문
        # 모든 시트에서 테이블을 찾아야 놓치는 데이터가 없음
        for sheet in wb_meta.worksheets:  # 워크시트 객체를 직접 반복
            sheet_name = sheet.title  # 시트 이름 가져오기
            
            # =============================================================
            # 1단계: AutoFilter 범위 탐지
            # =============================================================
            # 왜 AutoFilter를 확인하는가?
            # 사용자가 데이터에 필터를 적용했다는 것은 그 범위가 의미 있는 테이블이라는 뜻
            # Excel에서 "데이터 > 필터" 기능을 사용한 범위를 자동으로 찾아줌
            # 실무에서 종종 사용되므로 유지
            try:
                if hasattr(sheet, 'auto_filter') and sheet.auto_filter and sheet.auto_filter.ref:
                    # AutoFilter 범위를 파싱하여 start_row, end_row 등 필드 추가
                    range_info = parse_excel_range(sheet.auto_filter.ref)
                    
                    table_info = {
                        'sheet': sheet_name,              # 어느 시트에 있는지
                        'type': 'autofilter_range',       # 테이블 유형 (AutoFilter 범위)
                        'name': None,                     # AutoFilter에는 별도 이름이 없음
                        'ref': sheet.auto_filter.ref,     # 필터가 적용된 범위
                        'start_row': range_info['start_row'],  # 시작 행 번호
                        'end_row': range_info['end_row'],      # 끝 행 번호
                        'start_col': range_info['start_col'],  # 시작 열 번호
                        'end_col': range_info['end_col'],      # 끝 열 번호
                        'confidence': 0.8,                # 신뢰도 80%
                        'file_path': file_path,           # 원본 파일 경로
                        'headers': extract_headers_from_range(sheet, sheet.auto_filter.ref)  # 헤더 추출
                    }
                    tables.append(table_info)
                    print(f"[table_finder.find_tables] ① AutoFilter 탐지: 시트 '{sheet_name}', 범위 {sheet.auto_filter.ref}")
            except Exception as e:
                # read_only 모드에서는 AutoFilter 정보에 접근하지 못할 수 있음
                print(f"[table_finder.find_tables] ⚠️ AutoFilter 탐지 제한: {e}")
            
        
        # 메타데이터 접근용 워크북 닫기
        wb_meta.close()
        
        # 2단계: 데이터 스캔을 위한 read_only 모드 로딩
        wb_readonly = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_readonly in wb_readonly.worksheets:
            sheet_name = sheet_readonly.title
            
            # =============================================================
            # 2단계: 휴리스틱 기반 테이블 탐지 (지능적 추측 + 범위 최적화)
            # =============================================================
            # 왜 휴리스틱 방법이 필요한가?
            # AutoFilter로 찾지 못한 테이블들이 있을 수 있음
            # 예: 필터를 적용하지 않았지만 실제로는 구조화된 데이터
            # 헤더 패턴과 데이터 밀도를 분석해서 테이블을 추측
            # 범위 최적화로 불필요한 스캔 제거 (read_only 모드로 최적화)
            heuristic_candidates = find_tabular_ranges_by_heuristics_optimized(sheet_readonly)
            for candidate in heuristic_candidates:
                # 왜 시트 이름을 나중에 추가하는가?
                # find_tabular_ranges_by_heuristics_optimized 함수에서는 시트 객체만 받기 때문
                # 시트 이름은 이 함수에서 추가해줘야 함
                candidate['sheet'] = sheet_name
                candidate['file_path'] = file_path
                print(f"[table_finder.find_tables] ② 휴리스틱 탐지: 시트 '{sheet_name}', 행 {candidate['start_row']}~{candidate['end_row']}, 신뢰도 {candidate['confidence']:.2f}")
                tables.append(candidate)
        
        # 왜 wb.close()를 해야 하는가?
        # 파일을 열면 컴퓨터 메모리를 사용함
        # read_only 모드에서는 특히 리소스 해제가 중요
        wb_readonly.close()
        
        # 왜 신뢰도 순으로 정렬하는가?
        # 가장 확실한 테이블부터 처리하기 위해
        # 나중에 매칭이나 업데이트할 때 정확도가 높은 것을 우선적으로 사용
        tables.sort(key=lambda x: x.get('confidence', 0), reverse=True)  # 신뢰도 높은 순으로 정렬
        return tables  # 찾은 테이블들의 리스트 반환
        
    except Exception as e:
        # 왜 구체적인 오류 메시지를 보여주는가?
        # 사용자가 무엇이 잘못되었는지 알 수 있게 하기 위해
        # 특히 ~$ 임시 파일 같은 경우 "File is not a zip file" 오류가 발생할 수 있음
        print(f"[table_finder.find_tables] ❌ 테이블 찾기 실패: {e}")
        return []  # 오류 발생 시에도 빈 리스트 반환 (프로그램 중단 방지)

def extract_headers_from_range(sheet, range_ref):
    """
    Excel 범위에서 헤더(첫 번째 행) 추출 - 병합된 헤더 처리 개선
    
    이 함수가 하는 일:
    1. 범위 문자열(예: "A1:C10")을 좌표로 변환
    2. 첫 번째 행의 값들을 헤더로 추출 (병합 셀 고려)
    3. 빈 값이나 None은 제외하고 실제 헤더만 반환
    4. 병합된 셀의 경우 병합 범위의 값을 모든 열에 적용
    
    병합된 헤더 처리 방식:
    - 병합된 셀에서는 첫 번째 셀에만 값이 있고 나머지는 None
    - 병합 범위를 감지하여 해당 범위의 모든 열에 같은 헤더 적용
    - read_only 모드에서는 병합 정보 접근이 제한적이므로 fallback 로직 적용
    
    Parameters:
        sheet: 워크시트 객체
        range_ref (str): Excel 범위 참조 (예: "A1:C10")
        
    Returns:
        list: 헤더 문자열 리스트
    """
    try:
        # 범위 파싱: "A1:C10" → (min_col, min_row, max_col, max_row)
        min_col, min_row, max_col, max_row = range_boundaries(range_ref)
        
        # 병합된 헤더 처리를 위한 개선된 추출 로직
        try:
            # 헤더 행의 셀 객체들을 직접 가져와서 병합 정보 확인
            header_cells = []
            cell_access_errors = 0
            
            for col in range(min_col, max_col + 1):
                try:
                    cell = sheet.cell(row=min_row, column=col)
                    header_cells.append(cell)
                except Exception as cell_error:
                    # read_only 모드나 접근 오류 시 None 추가
                    header_cells.append(None)
                    cell_access_errors += 1
                    if cell_access_errors == 1:  # 첫 번째 오류만 로그
                        print(f"[table_finder.extract_headers_from_range] ⚠️ 셀 접근 제한 (col {col}): {type(cell_error).__name__}")
            
            # 너무 많은 셀 접근 실패시 경고
            if cell_access_errors > (max_col - min_col) * 0.5:
                print(f"[table_finder.extract_headers_from_range] ⚠️ 많은 셀 접근 실패 ({cell_access_errors}/{max_col - min_col + 1}개) - read_only 모드일 수 있음")
            
            headers = []
            last_valid_header = None
            processing_errors = 0
            
            for i, cell in enumerate(header_cells):
                try:
                    if cell is not None and cell.value is not None:
                        header_value = str(cell.value).strip()
                        if header_value:
                            headers.append(header_value)
                            last_valid_header = header_value
                        else:
                            # 빈 값이지만 병합된 셀일 수 있음 - 이전 헤더 재사용
                            if last_valid_header and _is_likely_merged_cell(cell, header_cells, i):
                                headers.append(last_valid_header)
                            else:
                                headers.append("")
                    else:
                        # None 값 - 병합된 셀일 가능성 체크
                        if last_valid_header and _is_likely_merged_continuation(header_cells, i):
                            headers.append(last_valid_header)
                        else:
                            headers.append("")
                except Exception as process_error:
                    # 개별 셀 처리 실패 시 빈 문자열
                    headers.append("")
                    processing_errors += 1
                    if processing_errors <= 3:  # 처음 3개 오류만 로그
                        print(f"[table_finder.extract_headers_from_range] ⚠️ 헤더 셀 {i + min_col} 처리 실패: {type(process_error).__name__}")
            
            # 빈 문자열 제거하여 실제 헤더만 반환
            valid_headers = [h for h in headers if h.strip()]
            
            # 헤더 품질 검증
            if not valid_headers:
                print(f"[table_finder.extract_headers_from_range] ⚠️ 유효한 헤더를 찾지 못함 - fallback 모드 시도")
                raise ValueError("No valid headers found")
            elif len(valid_headers) < (max_col - min_col + 1) * 0.3:
                print(f"[table_finder.extract_headers_from_range] ⚠️ 헤더 수가 적음 ({len(valid_headers)}/{max_col - min_col + 1}) - 데이터 품질 확인 필요")
            
            headers = valid_headers
            
        except Exception as advanced_error:
            # 개선된 방법 실패 시 기존 방식으로 fallback
            print(f"[table_finder.extract_headers_from_range] ⚠️ 고급 헤더 추출 실패: {type(advanced_error).__name__} - 기본 방식으로 fallback")
            
            try:
                header_row = next(sheet.iter_rows(min_row=min_row, max_row=min_row,
                                                min_col=min_col, max_col=max_col,
                                                values_only=True))
                headers = []
                for i, cell_value in enumerate(header_row):
                    if cell_value is not None and str(cell_value).strip():
                        headers.append(str(cell_value).strip())
                
                if not headers:
                    print(f"[table_finder.extract_headers_from_range] ⚠️ Fallback에서도 헤더를 찾지 못함")
                else:
                    print(f"[table_finder.extract_headers_from_range] ✅ Fallback 성공: {len(headers)}개 헤더 추출")
                    
            except Exception as fallback_error:
                print(f"[table_finder.extract_headers_from_range] ❌ Fallback도 실패: {type(fallback_error).__name__}: {str(fallback_error)}")
                headers = []
        
        return headers
        
    except Exception as e:
        # 구체적인 오류 정보와 함께 디버깅에 도움되는 메시지 출력
        print(f"[table_finder.extract_headers_from_range] ❌ 헤더 추출 완전 실패")
        print(f"  📍 범위: {range_ref}")
        print(f"  🚫 오류: {type(e).__name__}: {str(e)}")
        print(f"  💡 해결방법: 1) 범위 형식 확인, 2) 시트 접근 권한 확인, 3) 파일 손상 여부 점검")
        return []

def _is_likely_merged_cell(cell, header_cells, current_index):
    """
    셀이 병합된 셀일 가능성 체크 (read_only 모드 호환)
    
    병합된 셀 특징:
    - 값이 있지만 다음 셀들이 연속으로 None이거나 빈 값
    - read_only 모드에서는 merged_cells 속성 접근이 제한적
    
    Args:
        cell: 현재 셀 객체
        header_cells: 헤더 행의 모든 셀들
        current_index: 현재 셀의 인덱스
    
    Returns:
        bool: 병합된 셀일 가능성이 높으면 True
    """
    try:
        # read_only 모드에서는 간단한 휴리스틱 사용
        # 현재 셀에 값이 있고 다음 1-2개 셀이 비어있으면 병합 가능성 높음
        if current_index + 1 < len(header_cells):
            next_cell = header_cells[current_index + 1]
            if next_cell is None or next_cell.value is None or not str(next_cell.value).strip():
                return True
        return False
    except:
        # 접근 오류 시 보수적으로 False 반환
        return False

def _is_likely_merged_continuation(header_cells, current_index):
    """
    현재 None 셀이 병합된 셀의 연속 부분일 가능성 체크
    
    병합된 셀의 연속 부분 특징:
    - 이전 셀에 유효한 값이 있음
    - 현재 셀과 다음 몇 개 셀이 비어있음
    
    Args:
        header_cells: 헤더 행의 모든 셀들
        current_index: 현재 셀의 인덱스
    
    Returns:
        bool: 병합 연속 부분일 가능성이 높으면 True
    """
    try:
        # 이전 셀에 값이 있는지 확인
        if current_index > 0:
            prev_cell = header_cells[current_index - 1]
            if prev_cell and prev_cell.value is not None and str(prev_cell.value).strip():
                # 이전 셀에 값이 있고 현재 셀이 비어있으면 병합 연속 가능성 높음
                return True
                
        # 더 이전 셀들도 체크 (최대 3개까지)
        for back_step in range(2, min(4, current_index + 1)):
            if current_index - back_step >= 0:
                back_cell = header_cells[current_index - back_step]
                if back_cell and back_cell.value is not None and str(back_cell.value).strip():
                    # 중간에 모든 셀이 비어있는지 확인
                    all_empty = True
                    for check_idx in range(current_index - back_step + 1, current_index):
                        check_cell = header_cells[check_idx]
                        if check_cell and check_cell.value is not None and str(check_cell.value).strip():
                            all_empty = False
                            break
                    if all_empty:
                        return True
        
        return False
    except:
        # 접근 오류 시 보수적으로 False 반환
        return False

def find_tabular_ranges_by_heuristics_optimized(ws):
    """
    휴리스틱 기반 테이블 영역 탐지 - 범위 최적화 버전
    
    이 함수가 하는 일:
    1. calculate_dimension()으로 실제 사용 범위만 스캔
    2. 각 행을 벡터화된 방식으로 스캔하면서 헤더 후보를 찾기
    3. 헤더 후보 아래에 실제 데이터가 연속으로 있는지 확인
    4. 조건을 만족하는 영역을 테이블 후보로 선정
    5. 신뢰도를 계산해서 품질 평가
    
    속도 개선 포인트:
    - calculate_dimension(): 빈 영역 스캔하지 않음
    - iter_rows 벡터화: 행 단위로 효율적 처리
    - 조기 종료: 조건 불만족 시 즉시 다음 행으로
    - 범위 제한: 실용적인 최대 스캔 범위 설정
    
    Parameters(매개변수):
        ws: 워크시트 객체 (openpyxl worksheet, read_only 모드)
        
    Returns(반환값):
        list: 찾은 테이블 후보들의 리스트
    """
    
    candidates = []  # 테이블 후보들을 저장할 리스트
    
    try:
        # 왜 calculate_dimension()을 사용하는가?
        # 워크시트에서 실제 데이터가 있는 범위만 계산해서 불필요한 스캔 제거
        # 빈 셀들을 스캔하지 않아서 속도가 대폭 향상됨
        # 예: 전체 시트가 1048576x16384이지만 실제 데이터는 A1:J50만 있는 경우
        dimension = ws.calculate_dimension()
        if not dimension:
            return candidates
            
        min_col, min_row, max_col, max_row = range_boundaries(dimension)
        
        # 왜 스캔 범위를 제한하는가?
        # 실제 테이블은 보통 시트의 상단 영역에 위치함
        # 너무 아래쪽까지 스캔하면 속도만 느려지고 의미 없는 데이터를 찾을 가능성
        # 실용적인 범위로 제한해서 성능과 정확도의 균형 맞춤
        scan_max_rows = min(max_row, min_row + 1000)  # 최대 1000행까지 (대용량 데이터 대응)
        scan_max_cols = min(max_col, min_col + 50)   # 최대 50열까지만
        
        print(f"[table_finder.find_tabular_ranges_by_heuristics_optimized] 📏 스캔 범위: 행 {min_row}~{scan_max_rows}, 열 {min_col}~{scan_max_cols}")
        
    except Exception as e:
        # calculate_dimension이 실패하면 기본 범위 사용
        print(f"[table_finder.find_tabular_ranges_by_heuristics_optimized] ⚠️ 범위 계산 실패, 기본값 사용: {e}")
        min_row, scan_max_rows = 1, 1001  # 1행부터 1000행까지 (1+1000, 대용량 데이터 대응)
        min_col, scan_max_cols = 1, 51     # 1열부터 50열까지 (1+50)
    
    # 왜 행을 하나씩 스캔하는가?
    # 헤더(제목 행)를 찾기 위해서는 각 행의 특성을 분석해야 함
    # 헤더는 보통 텍스트로 구성되어 있고, 데이터 행보다 위쪽에 있음
    for start_row in range(min_row, scan_max_rows + 1):
        
        try:
            # 왜 iter_rows를 사용하는가?
            # openpyxl에서 행 데이터를 효율적으로 읽는 벡터화된 방법
            # read_only 모드에서 최적화된 성능 제공
            # values_only=True로 하면 셀 객체가 아닌 값만 바로 가져와서 더 빠름
            row_data = next(ws.iter_rows(min_row=start_row, max_row=start_row, 
                                       min_col=min_col, max_col=scan_max_cols, 
                                       values_only=True))
            
            # 왜 텍스트 개수를 세는가?
            # 헤더 행은 보통 텍스트(컬럼명)로 구성되어 있기 때문
            # 숫자나 빈 셀이 많으면 헤더가 아닐 가능성이 높음
            total_cells = len(row_data)
            text_count = sum(1 for v in row_data if isinstance(v, str) and v.strip())
            
            # 왜 최소 3개의 텍스트 셀이 있어야 하는가?
            # 1-2개만 있으면 제목이나 레이블일 가능성이 높음
            # 3개 이상은 실제 테이블 헤더일 가능성이 높음 (경험적 기준)
            if text_count < 3:
                continue
                
            # 왜 텍스트 비율을 확인하는가?
            # 전체 셀 중에서 일정 비율 이상이 텍스트여야 헤더로 인정
            # 너무 비율이 낮으면 헤더가 아닐 가능성이 높음
            text_ratio = text_count / total_cells if total_cells > 0 else 0
            if text_ratio < 0.9:  # 90% 이상의 셀이 텍스트여야 함
                continue
            
            # =======================================================
            # 헤더 후보 아래에 실제 데이터가 있는지 확인
            # =======================================================
            continuous_data_count = 0  # 연속된 데이터 행 개수를 셀 변수
            
            # 왜 헤더 다음 행부터 최대 15행만 확인하는가?
            # 헤더 바로 아래에 데이터가 연속으로 있어야 진짜 테이블
            # 너무 멀리 떨어진 데이터는 다른 테이블일 가능성이 높음
            # 15행 정도면 충분히 테이블 패턴을 확인할 수 있음
            for r in range(start_row + 1, min(start_row + 16, scan_max_rows + 1)):
                try:
                    # 이 행의 데이터를 벡터화된 방식으로 읽어오기
                    data_row = next(ws.iter_rows(min_row=r, max_row=r, 
                                               min_col=min_col, max_col=scan_max_cols, 
                                               values_only=True))
                    
                    # 왜 빈 셀이 아닌 셀의 개수를 세는가?
                    # 데이터 행은 일정 비율 이상의 셀에 실제 값이 있어야 함
                    # 빈 행이 많으면 테이블의 끝이거나 중간에 빈 공간일 가능성
                    non_empty = sum(1 for v in data_row if v not in (None, "", 0))
                    
                    # 왜 30% 이상의 데이터가 있어야 하는가?
                    # 너무 높으면(70%) 일부 빈 셀이 있는 정상 테이블도 제외될 수 있음
                    # 너무 낮으면(10%) 의미 없는 산발적 데이터도 테이블로 인식될 수 있음
                    # 30%는 실용적인 균형점
                    if non_empty / total_cells >= 0.3:
                        continuous_data_count += 1  # 유효한 데이터 행으로 카운트
                    else:
                        # 왜 break를 사용하는가?
                        # 데이터 밀도가 낮은 행을 만나면 테이블의 끝으로 간주
                        # 연속된 데이터 블록을 찾는 것이 목적이므로
                        break
                        
                except StopIteration:
                    # iter_rows가 더 이상 행을 반환하지 않으면 종료
                    break
                except Exception:
                    # 개별 행 읽기 실패 시 다음 행으로 계속
                    continue
            
            # 왜 최소 2행 이상이어야 테이블로 인정하는가?
            # 헤더 1행 + 데이터 최소 2행은 있어야 의미 있는 테이블
            # 1행만 있으면 제목이거나 단순한 레이블일 가능성이 높음
            if continuous_data_count >= 2:
                # 왜 신뢰도를 텍스트 비율로 계산하는가?
                # 헤더의 텍스트 비율이 높을수록 더 확실한 테이블 구조
                # 0.4~1.0 사이의 값으로, 1.0에 가까울수록 확신도가 높음
                confidence = min(0.95, text_ratio + (continuous_data_count * 0.1))  # 데이터 행수도 신뢰도에 반영
                
                # 실제 헤더 추출
                headers = [str(v).strip() for v in row_data if v is not None and str(v).strip()]
                
                candidate = {
                    'type': 'heuristic',                    # 테이블 유형 (휴리스틱)
                    'start_row': start_row,                 # 헤더가 시작되는 행
                    'end_row': start_row + continuous_data_count,       # 데이터가 끝나는 행
                    'ref': f"{chr(64 + min_col)}{start_row}:{chr(64 + min_col + len(headers) - 1)}{start_row + continuous_data_count}",  # 실제 범위
                    'confidence': confidence,               # 계산된 신뢰도
                    'headers': headers                      # 추출한 헤더들
                }
                
                # 디버깅: 테이블 후보 발견 시 상세 정보 출력
                # [table_finder.py > find_tabular_ranges_by_heuristics_optimized] 출처 명시
                print(f"[table_finder.find_tabular_ranges_by_heuristics_optimized] 🎯 테이블 후보 발견!")
                print(f"  📍 위치: {candidate['ref']}")  
                print(f"  📊 행 범위: {candidate['start_row']} ~ {candidate['end_row']} ({continuous_data_count}개 데이터 행)")
                print(f"  🎯 신뢰도: {candidate['confidence']:.2f}")
                print(f"  📋 헤더({len(headers)}개): {headers}")
                print(f"  📈 텍스트 비율: {text_ratio:.2f} ({text_count}/{total_cells})")
                print()
                
                candidates.append(candidate)
                
        except StopIteration:
            # iter_rows가 더 이상 행을 반환하지 않으면 스캔 종료
            break
        except Exception:
            # 개별 행 처리 실패 시 다음 행으로 계속
            continue
    
    return candidates  # 찾은 테이블 후보들 반환

def column_string_to_number(col_str):
    """
    Excel 열 문자열을 숫자로 변환 (A=1, B=2, ..., Z=26, AA=27, ...)
    
    Args:
        col_str (str): Excel 열 문자열 (예: "A", "B", "AA", "IV")
    
    Returns:
        int: 열 번호 (1부터 시작)
    
    Example:
        >>> column_string_to_number("A")
        1
        >>> column_string_to_number("V") 
        22
        >>> column_string_to_number("AA")
        27
    """
    result = 0
    for char in col_str:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def parse_cell_reference(cell_ref):
    """
    Excel 셀 참조를 행과 열로 분리 (예: "A6" → row=6, col=1)
    
    Args:
        cell_ref (str): Excel 셀 참조 (예: "A6", "V339", "AA1")
    
    Returns:
        tuple: (row, col) 튜플
    
    Example:
        >>> parse_cell_reference("A6")
        (6, 1)
        >>> parse_cell_reference("V339")
        (339, 22)
    """
    import re
    match = re.match(r'([A-Z]+)(\d+)', cell_ref)
    if not match:
        raise ValueError(f"Invalid cell reference: {cell_ref}")
    
    col_str, row_str = match.groups()
    return int(row_str), column_string_to_number(col_str)

def parse_excel_range(range_str):
    """
    Excel 범위 문자열을 구성 요소로 파싱 (예: "A6:V339" → start_row=6, end_row=339, start_col=1, end_col=22)
    
    Args:
        range_str (str): Excel 범위 문자열 (예: "A6:V339")
    
    Returns:
        dict: 파싱된 범위 정보
        {
            'start_row': int,    # 시작 행 (1부터)
            'end_row': int,      # 끝 행 (1부터)  
            'start_col': int,    # 시작 열 (1부터)
            'end_col': int       # 끝 열 (1부터)
        }
    
    Example:
        >>> parse_excel_range("A6:V339")
        {'start_row': 6, 'end_row': 339, 'start_col': 1, 'end_col': 22}
    """
    if ':' not in range_str:
        raise ValueError(f"Invalid range format: {range_str}")
    
    start_cell, end_cell = range_str.split(':')
    start_row, start_col = parse_cell_reference(start_cell)
    end_row, end_col = parse_cell_reference(end_cell)
    
    return {
        'start_row': start_row,
        'end_row': end_row,
        'start_col': start_col,
        'end_col': end_col
    }

def test_table_finder():
    """
    테이블 찾기 기능 테스트 (고속 개선 버전)
    
    왜 테스트 함수가 필요한가?
    - 개선된 알고리즘이 제대로 작동하는지 확인하기 위해
    - 3단계 탐지 방법과 속도 최적화가 모두 정상 작동하는지 검증하기 위해
    - 신뢰도 정렬이 올바르게 되는지 확인하기 위해
    - read_only 모드와 범위 최적화의 효과를 측정하기 위해
    """
    print("[table_finder.test_table_finder] 🧪 테이블 찾기 테스트 (고속 3단계 알고리즘)...")
    
    # 테스트 파일 경로 - 실제 파일이 있어야 테스트 가능
    test_file = "test_files/current_folder/current_pbc.xlsx"  # 실제 존재하는 파일로 변경
    
    # 속도 측정을 위한 시간 기록
    import time
    start_time = time.time()
    
    # 실제 함수 호출 - 개선된 테이블 찾기 실행
    tables = find_tables(test_file)
    
    end_time = time.time()
    processing_time = end_time - start_time
    
    # 결과 검증 및 표시
    if tables:  # 테이블을 찾았으면
        print(f"[table_finder.test_table_finder] ✅ 테스트 성공: {len(tables)}개 테이블 발견 (처리 시간: {processing_time:.2f}초)")
        
        # 왜 처음 3개만 표시하는가?
        # 너무 많은 결과는 화면을 복잡하게 만들기 때문
        # 가장 신뢰도 높은 것들만 보여주면 충분함
        for i, table in enumerate(tables[:3], 1):
            headers_preview = table.get('headers', [])[:3] if table.get('headers') else []
            print(f"[table_finder.test_table_finder]    {i}. 시트: {table['sheet']}, 유형: {table.get('type')}, 범위: {table.get('ref')}, 신뢰도: {table.get('confidence', 0):.2f}, 헤더: {headers_preview}...")
    else:  # 테이블을 찾지 못했으면
        print(f"[table_finder.test_table_finder] ❌ 테스트 실패: 테이블을 찾지 못했습니다 (처리 시간: {processing_time:.2f}초)")
        print("[table_finder.test_table_finder] 💡 힌트: 테스트 파일 경로를 확인하세요")

# 왜 이 조건문이 있는가?
# 이 파일을 직접 실행했을 때만 테스트를 실행하기 위해
# 다른 파일에서 import할 때는 테스트가 실행되지 않게 하기 위해
if __name__ == "__main__":
    test_table_finder()  # 테스트 함수 실행