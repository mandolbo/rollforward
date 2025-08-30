"""
Simple Worksheet Copy-Paste Module
워크시트 전체 복사-붙여넣기 (Ctrl+C, Ctrl+V 방식)
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

def copy_entire_worksheet(source_wb_path, source_ws_name, target_wb_path, target_ws_name):
    """
    워크시트 전체를 복사-붙여넣기 (Ctrl+C, Ctrl+V 방식)
    
    Args:
        source_wb_path: 소스 워크북 경로
        source_ws_name: 소스 워크시트 이름
        target_wb_path: 타겟 워크북 경로  
        target_ws_name: 타겟 워크시트 이름
    
    Returns:
        bool: 성공 여부
    """
    try:
        # 소스 및 타겟 워크북 로드
        source_wb = load_workbook(source_wb_path, data_only=False)
        target_wb = load_workbook(target_wb_path, data_only=False)
        
        source_ws = source_wb[source_ws_name]
        target_ws = target_wb[target_ws_name]
        
        # 타겟 워크시트 내용 모두 삭제 (워크시트는 유지)
        target_ws.delete_rows(1, target_ws.max_row)
        target_ws.delete_cols(1, target_ws.max_column)
        
        # 워크시트 전체 복사 (openpyxl의 간단한 방식)
        for row in source_ws.iter_rows():
            for cell in row:
                target_cell = target_ws.cell(row=cell.row, column=cell.column)
                
                # 값 복사
                target_cell.value = cell.value
                
                # 스타일 복사 (간단)
                if cell.has_style:
                    target_cell.font = cell.font.copy()
                    target_cell.border = cell.border.copy()
                    target_cell.fill = cell.fill.copy()
                    target_cell.number_format = cell.number_format
                    target_cell.alignment = cell.alignment.copy()
        
        # 열 너비 복사
        for col_letter in source_ws.column_dimensions:
            if col_letter in source_ws.column_dimensions:
                target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width
        
        # 행 높이 복사  
        for row_num in source_ws.row_dimensions:
            if row_num in source_ws.row_dimensions:
                target_ws.row_dimensions[row_num].height = source_ws.row_dimensions[row_num].height
        
        # 타겟 워크북 저장
        target_wb.save(target_wb_path)
        
        logger.info(f"워크시트 복사 완료: {source_ws_name} → {target_ws_name}")
        return True
        
    except Exception as e:
        logger.error(f"워크시트 복사 실패: {e}")
        return False

def simple_worksheet_replace(source_wb_path, source_ws_name, target_wb_path, target_ws_name):
    """
    더 간단한 워크시트 교체 방식 - 메모리 효율적
    
    Args:
        source_wb_path: 소스 워크북 경로
        source_ws_name: 소스 워크시트 이름
        target_wb_path: 타겟 워크북 경로
        target_ws_name: 타겟 워크시트 이름
    
    Returns:
        bool: 성공 여부
    """
    try:
        # 타겟 워크북 로드
        target_wb = load_workbook(target_wb_path)
        
        # 기존 워크시트 삭제 (참조 오류 방지를 위해 내용만 삭제)
        if target_ws_name in target_wb.sheetnames:
            target_ws = target_wb[target_ws_name]
            # 모든 셀 초기화
            for row in target_ws.iter_rows():
                for cell in row:
                    cell.value = None
                    cell.data_type = 's'
        
        # 소스 워크북에서 데이터 가져오기
        source_wb = load_workbook(source_wb_path, data_only=False)
        source_ws = source_wb[source_ws_name]
        
        # 워크시트 통째로 복사 (openpyxl의 내장 기능 활용)
        target_ws._cells = source_ws._cells.copy()
        target_ws.column_dimensions = source_ws.column_dimensions.copy()
        target_ws.row_dimensions = source_ws.row_dimensions.copy()
        target_ws.merged_cells = source_ws.merged_cells.copy()
        
        # 워크시트 속성 복사
        target_ws.sheet_format = source_ws.sheet_format
        target_ws.sheet_properties = source_ws.sheet_properties
        
        # 저장
        target_wb.save(target_wb_path)
        
        logger.info(f"간단한 워크시트 교체 완료: {source_ws_name} → {target_ws_name}")
        return True
        
    except Exception as e:
        logger.error(f"워크시트 교체 실패: {e}")
        return False

def ultra_simple_copy(source_wb_path, source_ws_name, target_wb_path, target_ws_name):
    """
    가장 간단한 복사 방식 - Ctrl+C, Ctrl+V와 유사
    """
    try:
        source_wb = load_workbook(source_wb_path)
        target_wb = load_workbook(target_wb_path)
        
        # 소스에서 타겟으로 직접 복사
        source_ws = source_wb[source_ws_name]
        target_ws = target_wb[target_ws_name]
        
        # 기존 데이터 클리어
        target_ws.delete_rows(1, target_ws.max_row)
        
        # 전체 데이터 복사 (가장 간단한 방식)
        for row in source_ws.values:
            target_ws.append(row)
            
        target_wb.save(target_wb_path)
        return True
        
    except Exception as e:
        logger.error(f"Ultra simple copy 실패: {e}")
        return False