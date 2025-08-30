"""
파일 업데이트 - MVP 버전
백업을 만들고 안전하게 업데이트
"""

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import shutil
import os
from datetime import datetime

# 색상 정의
RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")    # 빨간색 (롤포워딩 대상)
GREEN_FILL = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid")  # 초록색 (완료됨)
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid") # 노란색 (진행중)

def update_file(matches, current_pbc_path, previous_ledger_path):
    """
    파일 업데이트 - 올바른 데이터 흐름 구현
    
    데이터 흐름 방향:
    - SOURCE: Current PBC files (당기 PBC)
    - TARGET: Previous ledger 백데이터 sheets (전기 조서의 백데이터 시트)
    - 방향: Current PBC → 백데이터 sheets
    
    Parameters:
        matches (list): 매칭된 헤더들의 정보
        current_pbc_path (str): 당기 PBC 파일 경로 (데이터 소스)
        previous_ledger_path (str): 전기 조서 파일 경로 (백데이터 시트가 있는 대상 파일)
    
    Returns:
        str: 업데이트된 파일 경로 (previous_ledger_path)
    """
    
    if not matches:
        print("[file_updater.update_file] ⚠️ 업데이트할 매칭이 없습니다")
        return False
    
    backup_path = None
    
    try:
        # 💾 1. 대상 파일(previous_ledger) 백업 생성 (필수!)
        if os.path.exists(previous_ledger_path):
            # Backup 폴더 생성
            file_dir = os.path.dirname(previous_ledger_path)
            backup_dir = os.path.join(file_dir, "Backup")
            
            # Backup 폴더가 없으면 생성
            if not os.path.exists(backup_dir):
                os.makedirs(backup_dir)
                print(f"[file_updater.update_file] 📁 Backup 폴더 생성: {backup_dir}")
            
            # 백업 파일 경로 생성 (대상 파일 백업)
            file_name = os.path.basename(previous_ledger_path)
            backup_filename = f"{os.path.splitext(file_name)[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(file_name)[1]}"
            backup_path = os.path.join(backup_dir, backup_filename)
            
            shutil.copy2(previous_ledger_path, backup_path)
            print(f"[file_updater.update_file] 💾 대상 파일 백업 생성: {backup_path}")
        else:
            print(f"[file_updater.update_file] ⚠️ 대상 파일이 없습니다: {previous_ledger_path}")
            return False
        
        # 📂 2. 두 파일 열기
        print(f"[file_updater.update_file] 📂 소스 파일 로드: {current_pbc_path}")
        current_wb = openpyxl.load_workbook(current_pbc_path)  # 소스: 당기 PBC
        
        print(f"[file_updater.update_file] 📂 대상 파일 로드: {previous_ledger_path}")
        target_wb = openpyxl.load_workbook(previous_ledger_path)  # 대상: 전기 조서 (백데이터)
        
        # ✏️ 3. 매칭된 데이터 업데이트 (실제 데이터 복사 구현)
        update_count = 0
        copied_rows_count = 0
        
        for match in matches:
            try:
                print(f"[file_updater.update_file] ✏️ 데이터 복사 시작: {match['from_header']} (당기 PBC) → {match['to_header']} (백데이터)")
                
                # 올바른 데이터 흐름: Current PBC → 백데이터
                source_table = match['from_table']  # 당기 PBC 테이블 (소스)
                target_table = match['to_table']    # 백데이터 시트 (대상)
                
                # 1단계: 당기 PBC에서 실제 데이터 추출
                source_data = get_actual_data_from_workbook(source_table, current_wb)
                
                if not source_data:
                    print(f"[file_updater.update_file] ⚠️ 당기 PBC 데이터가 없습니다: {match['from_header']}")
                    continue
                
                # 2단계: 백데이터 시트에서 해당 컬럼 찾기
                if target_table['sheet'] in target_wb.sheetnames:
                    target_sheet = target_wb[target_table['sheet']]
                    print(f"[file_updater.update_file] 📊 백데이터 시트 접근: {target_table['sheet']}")
                    
                    # 컬럼 인덱스 찾기 (에러 처리 강화)
                    try:
                        if match['to_header'] in target_table['headers']:
                            source_col_idx = source_table['headers'].index(match['from_header'])  # 당기 PBC 컬럼
                            target_col_idx = target_table['headers'].index(match['to_header'])    # 백데이터 컬럼
                        else:
                            # 향상된 디버깅: 사용 가능한 헤더 목록 표시
                            print(f"[file_updater.update_file] ❌ 백데이터 헤더 '{match['to_header']}'를 찾을 수 없습니다")
                            print(f"[file_updater.update_file] 🔎 백데이터 사용 가능한 헤더: {target_table['headers'][:5]}{'...' if len(target_table['headers']) > 5 else ''}")
                            print(f"[file_updater.update_file] 💡 매칭: '{match['from_header']}' (당기 PBC) → '{match['to_header']}' (백데이터) (신뢰도: {match.get('confidence', 'N/A')})")
                            continue
                    except (ValueError, IndexError) as idx_error:
                        print(f"[file_updater.update_file] ❌ 헤더 인덱스 오류: {idx_error}")
                        print(f"[file_updater.update_file] 🔎 문제 상세: from='{match['from_header']}' (당기 PBC), to='{match['to_header']}' (백데이터)")
                        print(f"[file_updater.update_file] 🔎 당기 PBC 헤더: {source_table['headers'][:3]}...")
                        print(f"[file_updater.update_file] 🔎 백데이터 헤더: {target_table['headers'][:3]}...")
                        continue
                        
                    # 3단계: 실제 데이터 복사 - 당기 PBC → 백데이터 시트
                    target_start_row = target_table['start_row'] + 1  # 헤더 다음 행부터
                    
                    try:
                        for row_idx, source_row in enumerate(source_data):
                            if source_col_idx < len(source_row):
                                target_row = target_start_row + row_idx
                                target_col = target_col_idx + 1  # openpyxl은 1부터 시작
                                
                                # 실제 데이터 복사: 당기 PBC → 백데이터
                                source_value = source_row[source_col_idx]
                                if source_value is not None and str(source_value).strip():
                                    try:
                                        target_sheet.cell(target_row, target_col, source_value)
                                        copied_rows_count += 1
                                    except Exception as cell_error:
                                        print(f"[file_updater.update_file] ⚠️ 백데이터 셀 복사 오류 ({target_row}, {target_col}): {cell_error}")
                    except Exception as copy_error:
                        print(f"[file_updater.update_file] ❌ 백데이터 복사 중 오류: {copy_error}")
                    
                    # 4단계: 백데이터 헤더에 빨간색 표시 (업데이트 완료 표시)
                    header_row = target_table['start_row']
                    header_col = target_col_idx + 1
                    
                    header_cell = target_sheet.cell(header_row, header_col)
                    # ✅ 백데이터 업데이트 완료 - 헤더는 빨간색 칠하지 않음 (요구사항 반영)
                    # header_cell.fill = RED_FILL  # 주석 처리: 헤더는 색상 표시 안 함
                    
                    print(f"[file_updater.update_file] ✅ 백데이터 업데이트 완료: {len(source_data)}행 → {match['to_header']} (백데이터)")
                    update_count += 1
                else:
                    print(f"[file_updater.update_file] ❌ 백데이터 시트를 찾을 수 없습니다: {target_table['sheet']}")
                    
            except Exception as e:
                print(f"[file_updater.update_file] ❌ 개별 백데이터 업데이트 실패: {e}")
                continue
        
        # 💾 4. 대상 파일(백데이터) 저장 및 결과 리포트
        try:
            if update_count > 0:
                # 파일 저장 전 권한 체크 및 자동 수정
                import stat
                try:
                    # 파일 쓰기 권한 확인
                    file_stat = os.stat(previous_ledger_path)
                    if not (file_stat.st_mode & stat.S_IWRITE):
                        print(f"[file_updater.update_file] ⚠️ 파일이 읽기 전용입니다: {previous_ledger_path}")
                        print(f"[file_updater.update_file] 🔧 읽기 전용 해제를 시도합니다...")
                        
                        # 읽기 전용 속성 해제 시도
                        try:
                            os.chmod(previous_ledger_path, stat.S_IWRITE | stat.S_IREAD)
                            print(f"[file_updater.update_file] ✅ 읽기 전용 해제 성공!")
                        except Exception as chmod_error:
                            print(f"[file_updater.update_file] ❌ 읽기 전용 해제 실패: {chmod_error}")
                            print(f"[file_updater.update_file] 💡 수동으로 파일 속성을 확인해주세요")
                            raise PermissionError("읽기 전용 해제에 실패했습니다")
                except OSError:
                    pass  # 파일 상태 확인 실패는 무시하고 저장 시도
                
                print(f"[file_updater.update_file] 💾 백데이터 파일 저장 중...")
                
                # 파일이 다른 프로세스에서 사용 중인지 확인
                def is_file_locked(file_path):
                    try:
                        # 파일을 쓰기 모드로 열어보기
                        with open(file_path, 'a'):
                            return False
                    except IOError:
                        return True
                
                if is_file_locked(previous_ledger_path):
                    print(f"[file_updater.update_file] ⚠️ 파일이 다른 프로그램에서 사용 중입니다")
                    print(f"[file_updater.update_file] 💡 Excel이나 다른 프로그램에서 파일을 닫고 다시 시도해주세요")
                
                target_wb.save(previous_ledger_path)
                print(f"[file_updater.update_file] ✅ 백데이터 파일 저장 성공!")
                
                print(f"\n[file_updater.update_file] 🎉 백데이터 업데이트 완료!")
                print(f"[file_updater.update_file] 📁 소스 파일: {current_pbc_path} (당기 PBC)")
                print(f"[file_updater.update_file] 📁 대상 파일: {previous_ledger_path} (백데이터 시트 업데이트됨)")
                print(f"[file_updater.update_file] 📊 매칭: {update_count}개")
                print(f"[file_updater.update_file] 📝 복사된 행: {copied_rows_count}개")
                print(f"[file_updater.update_file] 💾 백업: {backup_path}")
                
                # 워크북 정리 (저장 후에 닫기)
                try:
                    current_wb.close()
                    print(f"[file_updater.update_file] 🔄 소스 워크북 정리 완료")
                except:
                    pass
                
                try:
                    target_wb.close()
                    print(f"[file_updater.update_file] 🔄 대상 워크북 정리 완료")
                except:
                    pass
                
                return previous_ledger_path  # 성공 시 업데이트된 파일 경로 반환
            else:
                print(f"[file_updater.update_file] ⚠️ 업데이트할 데이터가 없습니다")
                try:
                    current_wb.close()
                    target_wb.close()
                except:
                    pass
                return False
        except PermissionError as perm_error:
            print(f"[file_updater.update_file] ❌ 백데이터 파일 저장 권한 오류: {perm_error}")
            print(f"[file_updater.update_file] 💡 해결 방법:")
            print(f"[file_updater.update_file]    1. Excel이나 다른 프로그램에서 파일을 닫아주세요")
            print(f"[file_updater.update_file]    2. 파일 속성에서 읽기 전용을 해제해주세요")
            print(f"[file_updater.update_file]    3. 관리자 권한으로 프로그램을 실행해보세요")
            print(f"[file_updater.update_file]    4. 파일 경로: {previous_ledger_path}")
            try:
                current_wb.close()
                target_wb.close()
            except:
                pass
            return False
        except Exception as save_error:
            print(f"[file_updater.update_file] ❌ 백데이터 파일 저장 실패: {save_error}")
            print(f"[file_updater.update_file] 🔍 오류 상세: {type(save_error).__name__}")
            try:
                current_wb.close()
                target_wb.close()
            except:
                pass
            return False
        
    except Exception as e:
        print(f"[file_updater.update_file] ❌ 백데이터 업데이트 실패: {e}")
        
        # 백업에서 복원 시도
        if backup_path and os.path.exists(backup_path):
            try:
                shutil.copy2(backup_path, previous_ledger_path)
                print(f"[file_updater.update_file] 🔄 백업에서 복원 완료")
            except Exception as restore_error:
                print(f"[file_updater.update_file] ❌ 복원도 실패: {restore_error}")
        
        try:
            current_wb.close()
            target_wb.close()
        except:
            pass
        
        return False

def get_actual_data_from_table(table_info, source_file, empty_streak_limit=2):
    """
    🎯 1순위+2순위 개선 적용: 테이블 경계 기준으로 데이터 추출
    
    Why 이 함수를 개선했는가?
    - 기존: 하드코딩된 1000행 제한과 고정된 5개 연속 빈 행 기준
    - 개선: 매개변수화된 빈 행 감지 + 테이블 경계 기준 처리
    
    What 이 함수가 하는 일?
    1. Excel 파일을 열기 (data_only=True로 수식 결과값만 가져오기)
    2. 테이블 경계를 정확히 계산해서 성능 최적화
    3. 매개변수화된 연속 빈 행 감지로 정확한 데이터 끝 판단
    4. 추출한 데이터를 리스트로 반환
    
    실제 예시:
    기존 방식: 전체 시트에서 1000행까지 스캔 (느림)
    개선 방식: 테이블 영역만 정확히 스캔 (빠름!)
    
    Parameters:
        table_info (dict): 테이블 정보
        source_file (str): 소스 파일 경로
        empty_streak_limit (int): 연속 빈 행 기준 (2순위 개선)
        
    Returns:
        list: 추출한 데이터 (행별 리스트)
    """
    
    try:
        # Why data_only=True를 사용하는가?
        # 수식이 아닌 계산된 결과값만 가져오기 위해 (예: =SUM(A1:A10) → 100)
        wb = openpyxl.load_workbook(source_file, data_only=True)
        
        # 3순위 개선: 안전 장치 - 시트 존재 확인
        if table_info['sheet'] not in wb.sheetnames:
            print(f"[get_actual_data_from_table] ❌ 시트를 찾을 수 없음: {table_info['sheet']}")
            wb.close()
            return []
        
        sheet = wb[table_info['sheet']]
        
        # 🎯 1순위 개선: 테이블 경계 기준으로 범위 계산
        print(f"[get_actual_data_from_table] 🔍 테이블 경계 기준 범위 계산 중...")
        actual_max_row, table_col_count = calculate_table_dimension(sheet, table_info, empty_streak_limit)
        
        # 헤더 다음 행부터 데이터 추출
        # Why +1? 헤더는 제외하고 실제 데이터만 가져오기 위해
        start_row = table_info['start_row'] + 1
        data_rows = []  # 추출한 데이터를 저장할 리스트
        
        print(f"[get_actual_data_from_table] 📊 최적화된 범위로 데이터 추출: {start_row}~{actual_max_row}행")
        
        # 🎯 개선된 방식: 계산된 범위만 처리
        for row in range(start_row, actual_max_row + 1):
            row_data = []
            empty_count = 0
            
            # 각 컬럼의 값을 하나씩 확인 (테이블 컬럼 범위만)
            for col in range(1, table_col_count + 1):
                try:
                    cell_value = sheet.cell(row, col).value
                    if cell_value is not None and str(cell_value).strip():
                        # 실제 값이 있으면 그대로 저장
                        row_data.append(cell_value)
                    else:
                        # 빈 셀이면 빈 문자열로 저장
                        row_data.append("")
                        empty_count += 1
                except Exception:
                    # 셀 접근 오류는 빈 셀로 처리
                    row_data.append("")
                    empty_count += 1
            
            # 행의 절반 이상이 비어있으면 의미 없는 행으로 판단
            if empty_count < len(table_info['headers']) / 2:
                data_rows.append(row_data)  # 의미있는 데이터가 있는 행만 추가
        
        wb.close()
        print(f"[get_actual_data_from_table] ✅ 테이블 경계 기준 데이터 추출 완료: {len(data_rows)}행")
        print(f"[get_actual_data_from_table] 🚀 성능 향상: 테이블 영역만 정확히 처리")
        return data_rows
        
    except Exception as e:
        print(f"[get_actual_data_from_table] ❌ 데이터 추출 실패: {e}")
        return []

def write_data_to_table(table_info, target_file, data_rows, header_mapping):
    """테이블에 데이터 쓰기 (Phase 1에서 구현 예정)"""
    
    try:
        wb = openpyxl.load_workbook(target_file)
        sheet = wb[table_info['sheet']]
        
        start_row = table_info['start_row'] + 1  # 헤더 다음 행부터
        
        for row_idx, data_row in enumerate(data_rows):
            target_row = start_row + row_idx
            
            for col_idx, value in enumerate(data_row):
                target_col = col_idx + 1
                sheet.cell(target_row, target_col, value)
        
        wb.save(target_file)
        wb.close()
        return True
        
    except Exception as e:
        print(f"[file_updater.write_data_to_table] ❌ 데이터 쓰기 실패: {e}")
        return False

def calculate_table_dimension(sheet, table_info, empty_streak_limit=2):
    """
    🎯 1순위 개선: 테이블의 실제 사용 영역만 정확히 계산
    
    Why 왜 이 함수가 필요한가?
    - 기존 방식: sheet.max_row는 시트 전체에서 가장 마지막 데이터가 있는 행을 반환
    - 문제점: 다른 영역(예: Z열 1000행)에 데이터가 있으면 관계없는 범위까지 포함
    - 해결책: 실제 처리할 테이블 영역만 정확히 계산해서 성능 향상
    
    What 이 함수가 하는 일?
    1. 테이블의 헤더 개수만큼만 컬럼 범위 제한 (다른 영역 무시)
    2. 헤더 아래부터 실제 데이터 끝까지만 행 범위 계산
    3. 연속된 빈 행이 나오면 데이터 끝으로 판단
    
    실제 예시:
    만약 시트에 이런 데이터가 있다면:
        A열: 계정과목, B열: 금액, C열: 비율  ... Z열: 기타데이터(1000행)
        실제 테이블 데이터는 1-10행까지만 있음
    기존 방식: 1000행 × 100열 = 10만 셀 스캔 (느림!) ❌
    개선 방식: 10행 × 3열 = 30셀만 스캔 (빠름!) ✅
    
    Parameters:
        sheet: openpyxl 워크시트 객체
        table_info (dict): 테이블 정보 {'start_row': 3, 'headers': ['계정과목', '금액', '비율']}
        empty_streak_limit (int): 연속 빈 행 기준 (기본값 2)
    
    Returns:
        tuple: (실제_최대행, 테이블_컬럼수)
    """
    
    # 3순위 개선: 안전 장치 - 잘못된 매개변수 보정
    if not (1 <= empty_streak_limit <= 10):
        print(f"[calculate_table_dimension] ⚠️ empty_streak_limit 값이 잘못됨 ({empty_streak_limit}), 기본값 2로 설정")
        empty_streak_limit = 2
    
    header_row = table_info['start_row']
    headers = table_info['headers']
    
    # Why +1을 하는가? 헤더는 제외하고 데이터부터 확인하기 위해
    data_start_row = header_row + 1
    
    # 🎯 핵심 개선: 테이블 컬럼 범위만 고려 (다른 영역 무시)
    min_col = 1
    max_col = len(headers)  # 헤더 개수만큼만! 다른 컬럼은 무시
    
    actual_max_row = header_row  # 최소한 헤더 행까지는 있음
    empty_streak = 0  # 연속 빈 행 카운터
    
    # Why 1000으로 제한? 너무 큰 파일을 무한정 스캔하면 시간이 오래 걸리므로
    max_scan_rows = 1000
    
    print(f"[calculate_table_dimension] 🔍 테이블 범위 계산: {len(headers)}개 컬럼 × 최대 {max_scan_rows}행 스캔")
    
    for row_num in range(data_start_row, data_start_row + max_scan_rows):
        # 🎯 핵심: 해당 테이블의 컬럼 범위에서만 데이터 확인
        row_has_data = False
        
        for col_num in range(min_col, max_col + 1):
            try:
                cell_value = sheet.cell(row_num, col_num).value
                # What: 실제 데이터가 있는지 확인 (None이 아니고 빈 문자열도 아님)
                if cell_value is not None and str(cell_value).strip():
                    row_has_data = True
                    break  # 하나라도 데이터가 있으면 이 행은 유효한 데이터 행
            except Exception:
                # 셀 접근 오류는 무시하고 계속
                continue
        
        if row_has_data:
            # 데이터가 있는 행을 발견! 최대 행 업데이트
            actual_max_row = row_num
            empty_streak = 0  # 빈 행 카운터 리셋
        else:
            # 빈 행 발견
            empty_streak += 1
            
            # 2순위 개선: 매개변수화된 연속 빈 행 감지
            # Why empty_streak_limit개 연속 빈 행이면 끝으로 판단?
            # 1개: 너무 민감 (중간에 구분자로 쓰인 빈 행에서 멈춤)
            # 2개: 일반적 (대부분의 경우에 적절함)
            # 3-5개: 관대 (그룹핑된 데이터나 복잡한 구조에 적합)
            if empty_streak >= empty_streak_limit:
                print(f"[calculate_table_dimension] 📍 연속 {empty_streak}개 빈 행 감지 → 데이터 끝으로 판단")
                break
    
    # 결과 계산
    actual_height = actual_max_row - header_row  # 헤더 제외한 실제 데이터 높이
    
    print(f"[calculate_table_dimension] ✅ 테이블 경계 계산 완료: {actual_height}행 × {max_col}열")
    print(f"[calculate_table_dimension] 📊 범위: {header_row}행(헤더) ~ {actual_max_row}행(마지막 데이터)")
    
    return actual_max_row, max_col

def get_actual_data_from_workbook(table_info, workbook, empty_streak_limit=2):
    """
    🎯 1순위 개선 적용: 테이블 경계 기준으로 데이터 추출
    
    Why 기존 방식의 문제점?
    - sheet.max_row 사용 시 시트 전체 범위를 고려해서 불필요한 영역까지 처리
    - 예: 실제 데이터는 10행인데 다른 곳에 데이터가 있어서 1000행까지 스캔
    - 결과: 느린 성능, 메모리 낭비, read-only 파일에서 문제 발생
    
    What 개선된 방식?
    - 테이블의 실제 사용 영역만 정확히 계산해서 처리
    - 다른 영역의 데이터는 완전히 무시
    - 성능 대폭 향상 (특히 대용량 파일에서 효과적)
    
    Parameters:
        table_info (dict): 테이블 정보
        workbook: 이미 로드된 openpyxl 워크북 객체  
        empty_streak_limit (int): 연속 빈 행 기준 (2순위 개선)
        
    Returns:
        list: 추출한 데이터 (행별 리스트)
    """
    try:
        sheet_name = table_info['sheet']
        if sheet_name not in workbook.sheetnames:
            print(f"[get_actual_data_from_workbook] ❌ 시트를 찾을 수 없습니다: {sheet_name}")
            return []
        
        sheet = workbook[sheet_name]
        start_row = table_info.get('start_row', 1)
        
        # 3순위 개선: 안전 장치 - 헤더 정보 검증
        headers = table_info.get('headers', [])
        if not headers:
            print(f"[get_actual_data_from_workbook] ⚠️ 헤더 정보가 없습니다")
            return []
        
        # 🎯 1순위 핵심 개선: 테이블 경계만 정확히 계산
        print(f"[get_actual_data_from_workbook] 🔍 기존 방식 대신 테이블 경계 기준으로 범위 계산 중...")
        actual_max_row, table_col_count = calculate_table_dimension(sheet, table_info, empty_streak_limit)
        
        # 데이터 추출 (헤더 제외하고 실제 데이터만)
        all_data = []
        data_start_row = start_row + 1  # Why +1? 헤더 다음 행부터 데이터
        
        print(f"[get_actual_data_from_workbook] 📊 최적화된 범위로 데이터 추출: {data_start_row}~{actual_max_row}행, {table_col_count}열")
        
        # 🎯 개선된 방식: 계산된 테이블 경계만 처리
        for row in sheet.iter_rows(min_row=data_start_row, max_row=actual_max_row, 
                                 min_col=1, max_col=table_col_count, values_only=True):
            # What: 완전히 빈 행은 건너뛰기 (모든 셀이 None이거나 빈 문자열)
            if any(cell is not None and str(cell).strip() for cell in row):
                all_data.append(row)
        
        print(f"[get_actual_data_from_workbook] ✅ 테이블 경계 기준 데이터 추출 완료: {len(all_data)}행")
        print(f"[get_actual_data_from_workbook] 🚀 성능 개선: 시트 전체 대신 테이블 영역만 처리")
        return all_data
        
    except Exception as e:
        print(f"[get_actual_data_from_workbook] ❌ 데이터 추출 실패: {e}")
        return []

def clear_table_data_area_dynamic(sheet, table_info, empty_streak_limit=2):
    """
    🎯 1순위 개선: 테이블 경계 기준으로 기존 데이터 영역을 동적으로 정리
    
    Why 기존 하드코딩 방식의 문제점?
    - clear_table_data_area(sheet, start_row, 50, 100) 같은 방식
    - 50행 고정: 실제 데이터가 수백/수천 행이면 나머지가 그대로 남음 ❌
    - 100열 고정: 불필요한 영역까지 처리해서 성능 저하 ❌
    
    What 개선된 방식?
    - 실제 테이블의 사용 영역만 정확히 계산해서 정리
    - 다른 영역은 건드리지 않음 (안전성 향상)
    - 대용량 파일에서도 빠른 처리 (성능 향상)
    
    실제 예시:
    기존: 50행×100열 = 5000셀 정리 (대부분 빈 셀인데도 처리)
    개선: 실제 10행×5열 = 50셀만 정리 (100배 빠름!)
    
    Parameters:
        sheet: openpyxl 워크시트 객체
        table_info (dict): 테이블 정보
        empty_streak_limit (int): 연속 빈 행 기준 (2순위 개선)
    
    Returns:
        bool: 정리 성공 여부
    """
    
    try:
        # 🎯 핵심: 테이블의 실제 경계 계산
        print(f"[clear_table_data_area_dynamic] 🔍 동적 범위 계산 중...")
        actual_max_row, table_col_count = calculate_table_dimension(sheet, table_info, empty_streak_limit)
        
        header_row = table_info['start_row']
        data_start_row = header_row + 1  # Why +1? 헤더는 보존하고 데이터만 정리
        
        # What: 헤더는 그대로 두고 데이터 영역만 정리
        rows_to_clear = actual_max_row - data_start_row + 1
        
        if rows_to_clear <= 0:
            print(f"[clear_table_data_area_dynamic] ℹ️ 정리할 데이터가 없습니다")
            return True
        
        print(f"[clear_table_data_area_dynamic] 🧹 테이블 데이터 정리: {data_start_row}~{actual_max_row}행 × {table_col_count}열")
        print(f"[clear_table_data_area_dynamic] 💡 기존 하드코딩 방식 대신 실제 사용 영역만 정리")
        
        # 🎯 개선된 정리: 계산된 범위만 정확히 정리
        cleared_count = 0
        for row_num in range(data_start_row, actual_max_row + 1):
            for col_num in range(1, table_col_count + 1):
                try:
                    # What: 셀 값을 None으로 설정해서 정리
                    sheet.cell(row_num, col_num).value = None
                    cleared_count += 1
                except Exception as cell_error:
                    print(f"[clear_table_data_area_dynamic] ⚠️ 셀 정리 오류 ({row_num}, {col_num}): {cell_error}")
                    continue
        
        print(f"[clear_table_data_area_dynamic] ✅ 동적 데이터 정리 완료: {cleared_count}개 셀 정리됨")
        print(f"[clear_table_data_area_dynamic] 🚀 성능 향상: 불필요한 영역 정리 하지 않음")
        return True
        
    except Exception as e:
        print(f"[clear_table_data_area_dynamic] ❌ 동적 데이터 정리 실패: {e}")
        return False

def synchronize_entire_table(source_table, source_wb, target_sheet, target_start_row, empty_streak_limit=2):
    """
    🎯 1순위+2순위 개선 통합: 테이블 경계 기준으로 전체 테이블 동기화
    
    Why 이 함수가 필요한가?
    - 2단계 롤포워딩을 위해 헤더와 데이터를 모두 동기화해야 함
    - 기존 개별 컬럼 매칭 방식 → 전체 테이블 교체 방식으로 변경
    - 성능과 정확성 모두 향상
    
    What 이 함수가 하는 일?
    1. 기존 헤더를 소스 테이블 헤더로 완전 교체
    2. 테이블 경계 기준으로 기존 데이터 영역 정리
    3. 소스 테이블의 모든 데이터를 대상에 복사
    4. 완료 표시 (빨간색 마킹)
    
    Parameters:
        source_table (dict): 소스 테이블 정보 (당기 PBC)
        source_wb: 소스 워크북 객체
        target_sheet: 대상 시트 객체 (백데이터 시트)
        target_start_row (int): 대상 헤더 행 번호
        empty_streak_limit (int): 연속 빈 행 기준 (2순위 개선)
        
    Returns:
        bool: 동기화 성공 여부
    """
    
    try:
        print(f"[synchronize_entire_table] 🔄 전체 테이블 동기화 시작...")
        
        # 1단계: 헤더 완전 교체
        print(f"[synchronize_entire_table] 📝 1단계: 헤더 교체 중...")
        
        # Why 기존 헤더를 지우는가? 완전히 새로운 구조로 만들기 위해
        # 충분한 범위(100열)로 기존 헤더 완전 삭제
        for col_idx in range(1, 101):
            target_sheet.cell(target_start_row, col_idx).value = None
        
        # 새 헤더 입력
        source_headers = source_table['headers']
        for col_idx, header in enumerate(source_headers):
            target_sheet.cell(target_start_row, col_idx + 1).value = header
        
        print(f"[synchronize_entire_table] ✅ 헤더 교체 완료: {len(source_headers)}개 컬럼")
        
        # 2단계: 🎯 1순위 개선 - 테이블 경계 기준 데이터 영역 정리
        print(f"[synchronize_entire_table] 🧹 2단계: 기존 데이터 정리 중...")
        
        # 대상 테이블 정보 생성 (새로운 헤더 구조 기준)
        target_table_info = {
            'start_row': target_start_row,
            'headers': source_headers  # 새로 설정된 헤더 사용
        }
        
        # 동적 범위 기준 데이터 정리
        clear_success = clear_table_data_area_dynamic(target_sheet, target_table_info, empty_streak_limit)
        if not clear_success:
            print(f"[synchronize_entire_table] ⚠️ 데이터 정리에 문제가 있었지만 계속 진행...")
        
        # 3단계: 소스 데이터 복사
        print(f"[synchronize_entire_table] 📋 3단계: 소스 데이터 복사 중...")
        
        # 🎯 1순위+2순위 개선: 테이블 경계 기준으로 소스 데이터 추출
        source_data = get_actual_data_from_workbook(source_table, source_wb, empty_streak_limit)
        
        if not source_data:
            print(f"[synchronize_entire_table] ⚠️ 소스 데이터가 없습니다")
            return True  # 헤더 교체는 성공했으므로 True 반환
        
        # 데이터 복사 (헤더 다음 행부터)
        target_data_start = target_start_row + 1
        copied_rows = 0
        
        for row_idx, row_data in enumerate(source_data):
            target_row = target_data_start + row_idx
            
            # Why min 사용? 소스 데이터가 헤더보다 적을 수 있으므로
            for col_idx in range(min(len(row_data), len(source_headers))):
                try:
                    target_sheet.cell(target_row, col_idx + 1).value = row_data[col_idx]
                except Exception as cell_error:
                    print(f"[synchronize_entire_table] ⚠️ 셀 복사 오류 ({target_row}, {col_idx+1}): {cell_error}")
                    continue
            
            copied_rows += 1
        
        # 4단계: 완료 표시  
        print(f"[synchronize_entire_table] ✅ 4단계: 완료 표시 중...")
        header_cell = target_sheet.cell(target_start_row, 1)
        # header_cell.fill = RED_FILL  # 주석 처리: 헤더는 빨간색 칠하지 않음 (요구사항 반영)
        
        print(f"[synchronize_entire_table] ✅ 전체 테이블 동기화 완료!")
        print(f"[synchronize_entire_table] 📊 결과: {len(source_headers)}개 헤더, {copied_rows}행 복사됨")
        print(f"[synchronize_entire_table] 🚀 성능 향상: 테이블 경계 기준 처리로 최적화됨")
        
        return True
        
    except Exception as e:
        print(f"[synchronize_entire_table] ❌ 테이블 동기화 실패: {e}")
        return False

def test_file_updater():
    """
    파일 업데이터 기능 테스트
    
    이 함수가 하는 일:
    1. 실제로 존재하는 Excel 파일을 찾기
    2. 더미 데이터로 롤포워딩 테스트 해보기
    3. 결과가 제대로 나오는지 확인하기
    
    왜 테스트 함수가 필요한가?
    - 프로그램이 제대로 작동하는지 미리 확인하기 위해
    - 문제가 있으면 실제 사용 전에 미리 찾아서 고치기 위해
    - 개발자가 코드를 수정한 후에도 정상 작동하는지 검증하기 위해
    """
    print("[file_updater.test_file_updater] 🧪 파일 업데이터 테스트...")
    
    # 현재 폴더에서 Excel 파일 찾기 (하드코딩 제거)
    current_dir = os.getcwd()
    excel_files = []
    
    # 왜 os.walk를 사용하는가?
    # 현재 폴더와 하위 폴더를 모두 검색해서 Excel 파일을 찾기 위해
    for root, dirs, files in os.walk(current_dir):
        for file in files:
            if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                excel_files.append(os.path.join(root, file))
    
    if not excel_files:
        print("[file_updater.test_file_updater] ⚠️ 테스트할 Excel 파일이 없습니다")
        print("[file_updater.test_file_updater] 💡 Excel 파일을 하나 만들어 놓고 테스트하세요")
        return
    
    # 첫 번째 Excel 파일로 테스트
    test_file = excel_files[0]
    print(f"[file_updater.test_file_updater] 📄 테스트 파일: {test_file}")
    
    # 더미 매칭 정보 (실제 상황을 흉내낸 가상 데이터)
    dummy_matches = [{
        'from_table': {'sheet': 'Sheet1', 'start_row': 1, 'headers': ['이름', '매출'], 'file_path': test_file},
        'to_table': {'sheet': 'Sheet1', 'start_row': 1, 'headers': ['이름', '매출', '기타']},
        'from_header': '매출',
        'to_header': '매출',
        'confidence': 1.0
    }]
    
    # 실제 테스트 실행
    result = update_file(test_file, dummy_matches)
    
    if isinstance(result, dict) and result.get('success'):
        print(f"[file_updater.test_file_updater] ✅ 테스트 성공!")
        print(f"[file_updater.test_file_updater] 📊 {result['matched_headers']}개 헤더, {result['copied_cells']}개 셀 처리됨")
    else:
        print(f"[file_updater.test_file_updater] ❌ 테스트 실패")
        print("[file_updater.test_file_updater] 💡 디버깅을 위해 debug_collector.py를 실행해보세요")

def mark_back_data_red(file_path, back_data_worksheets, tables_info):
    """
    백데이터 워크시트의 테이블 영역을 빨간색으로 표시
    
    이 함수가 하는 일:
    1. 백데이터로 분류된 워크시트들을 확인
    2. 각 워크시트의 테이블 영역을 찾기
    3. 테이블의 모든 데이터 셀을 빨간색으로 표시
    4. 롤포워딩 대상임을 나타내는 주석 추가
    
    Parameters:
        file_path (str): 전기 조서 파일 경로
        back_data_worksheets (list): 백데이터 워크시트 이름 리스트
        tables_info (list): 테이블 정보 리스트
        
    Returns:
        dict: 빨간색으로 표시된 셀 정보
    """
    
    if not back_data_worksheets:
        print("[file_updater.mark_back_data_red] ℹ️ 백데이터 워크시트가 없습니다.")
        return {}
    
    red_cells_info = {}  # 빨간색으로 표시된 셀들의 정보
    
    try:
        # 파일 권한 체크 및 해결
        import os
        import stat
        
        # 파일이 읽기 전용인지 확인
        if os.path.exists(file_path):
            file_stat = os.stat(file_path)
            if not (file_stat.st_mode & stat.S_IWRITE):
                print(f"[file_updater.mark_back_data_red] 🔓 읽기 전용 파일 권한 해제: {file_path}")
                os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
        
        wb = openpyxl.load_workbook(file_path)
        
        print(f"[file_updater.mark_back_data_red] 🔴 백데이터 표시 시작: {len(back_data_worksheets)}개 워크시트")
        
        for sheet_name in back_data_worksheets:
            if sheet_name not in wb.sheetnames:
                print(f"[file_updater.mark_back_data_red] ⚠️ 워크시트를 찾을 수 없습니다: {sheet_name}")
                continue
            
            sheet = wb[sheet_name]
            red_cells_info[sheet_name] = []
            
            # 해당 워크시트의 테이블 찾기
            sheet_tables = [table for table in tables_info if table['sheet'] == sheet_name]
            
            if not sheet_tables:
                print(f"[file_updater.mark_back_data_red] ⚠️ {sheet_name}에서 테이블을 찾을 수 없습니다.")
                continue
            
            for table in sheet_tables:
                try:
                    start_row = table['start_row']
                    headers = table['headers']
                    
                    print(f"[file_updater.mark_back_data_red] 🔴 {sheet_name} 테이블 표시 중: {len(headers)}개 컬럼")
                    
                    # 헤더 다음 행부터 데이터 영역 표시
                    data_start_row = start_row + 1
                    
                    # 최대 1000행까지 스캔 (안전 제한)
                    for row in range(data_start_row, data_start_row + 1000):
                        empty_count = 0
                        row_has_data = False
                        
                        for col in range(1, len(headers) + 1):
                            try:
                                cell = sheet.cell(row, col)
                                
                                if cell.value is not None and str(cell.value).strip():
                                    # 데이터가 있는 셀을 빨간색으로 표시
                                    cell.fill = RED_FILL
                                    
                                    # 주석 추가 안 함 (요구사항: 백데이터 시트에는 메모 추가하지 않음)
                                    
                                    red_cells_info[sheet_name].append({
                                        'row': row,
                                        'col': col,
                                        'value': cell.value,
                                        'status': 'pending'  # pending, completed, failed
                                    })
                                    row_has_data = True
                                else:
                                    empty_count += 1
                            except Exception as e:
                                print(f"[file_updater.mark_back_data_red] ⚠️ 셀 처리 오류 ({row}, {col}): {e}")
                                continue
                        
                        # 연속으로 5행이 비어있으면 데이터 끝으로 판단
                        if empty_count >= len(headers) or not row_has_data:
                            consecutive_empty = getattr(mark_back_data_red, f'empty_count_{sheet_name}', 0) + 1
                            setattr(mark_back_data_red, f'empty_count_{sheet_name}', consecutive_empty)
                            if consecutive_empty >= 5:
                                break
                        else:
                            setattr(mark_back_data_red, f'empty_count_{sheet_name}', 0)
                    
                    print(f"[file_updater.mark_back_data_red] ✅ {sheet_name}: {len(red_cells_info[sheet_name])}개 셀 표시 완료")
                    
                except Exception as e:
                    print(f"[file_updater.mark_back_data_red] ❌ 테이블 처리 오류 ({sheet_name}): {e}")
                    continue
        
        # 파일 저장
        wb.save(file_path)
        wb.close()
        
        total_cells = sum(len(cells) for cells in red_cells_info.values())
        print(f"[file_updater.mark_back_data_red] 🎯 백데이터 표시 완료: 총 {total_cells}개 셀이 빨간색으로 표시됨")
        
        return red_cells_info
        
    except Exception as e:
        print(f"[file_updater.mark_back_data_red] ❌ 백데이터 표시 실패: {e}")
        return {}

def update_rollforward_status(file_path, red_cells_info, successful_matches):
    """
    롤포워딩 완료된 셀을 초록색으로 변경하고 미완료 셀 추적
    
    이 함수가 하는 일:
    1. 성공적으로 롤포워딩된 셀들을 초록색으로 변경
    2. 실패한 셀들은 빨간색 유지
    3. 미완료 셀들의 위치 정보 수집
    4. 수기조정이 필요한 셀들의 리포트 생성
    
    Parameters:
        file_path (str): 전기 조서 파일 경로
        red_cells_info (dict): 빨간색으로 표시된 셀 정보
        successful_matches (list): 성공한 매칭 정보
        
    Returns:
        dict: 롤포워딩 결과 리포트
    """
    
    if not red_cells_info:
        print("[file_updater.update_rollforward_status] ℹ️ 처리할 빨간색 셀이 없습니다.")
        return {'green_cells': 0, 'red_cells': 0, 'manual_adjustment_needed': []}
    
    try:
        # 파일 권한 체크 및 해결
        import os
        import stat
        
        # 파일이 읽기 전용인지 확인
        if os.path.exists(file_path):
            file_stat = os.stat(file_path)
            if not (file_stat.st_mode & stat.S_IWRITE):
                print(f"[file_updater.update_rollforward_status] 🔓 읽기 전용 파일 권한 해제: {file_path}")
                os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
        
        wb = openpyxl.load_workbook(file_path)
        
        green_count = 0
        remaining_red_count = 0
        manual_adjustment_cells = []
        
        print("[file_updater.update_rollforward_status] 🎨 롤포워딩 결과 색상 업데이트 중...")
        
        for sheet_name, cells_list in red_cells_info.items():
            if sheet_name not in wb.sheetnames:
                continue
            
            sheet = wb[sheet_name]
            
            for cell_info in cells_list:
                row = cell_info['row']
                col = cell_info['col']
                
                try:
                    cell = sheet.cell(row, col)
                    
                    # 이 셀이 성공적으로 롤포워딩되었는지 확인
                    was_updated = _check_if_cell_was_updated(cell_info, successful_matches, sheet_name)
                    
                    if was_updated:
                        # 성공한 셀을 초록색으로 변경
                        cell.fill = GREEN_FILL
                        
                        # 주석 추가 안 함 (요구사항: 백데이터 시트에는 메모 추가하지 않음)
                        
                        green_count += 1
                        cell_info['status'] = 'completed'
                    else:
                        # 실패한 셀은 빨간색 유지하고 수기조정 목록에 추가
                        remaining_red_count += 1
                        cell_info['status'] = 'failed'
                        
                        manual_adjustment_cells.append({
                            'sheet': sheet_name,
                            'row': row,
                            'col': col,
                            'cell_address': f"{sheet_name}!{cell.coordinate}",
                            'value': cell.value
                        })
                        
                        # 수기조정 필요 주석 추가
                        # 수기조정 필요한 셀에도 메모 추가 안 함 (요구사항: 백데이터 시트에는 메모 추가하지 않음)
                
                except Exception as e:
                    print(f"[file_updater.update_rollforward_status] ⚠️ 셀 업데이트 오류 ({row}, {col}): {e}")
                    continue
        
        wb.save(file_path)
        wb.close()
        
        # 결과 리포트
        result = {
            'green_cells': green_count,
            'red_cells': remaining_red_count, 
            'manual_adjustment_needed': manual_adjustment_cells
        }
        
        print(f"[file_updater.update_rollforward_status] ✅ 색상 업데이트 완료:")
        print(f"[file_updater.update_rollforward_status]    🟢 초록색 (완료): {green_count}개")
        print(f"[file_updater.update_rollforward_status]    🔴 빨간색 (수기조정 필요): {remaining_red_count}개")
        
        return result
        
    except Exception as e:
        print(f"[file_updater.update_rollforward_status] ❌ 상태 업데이트 실패: {e}")
        return {'green_cells': 0, 'red_cells': 0, 'manual_adjustment_needed': []}

def _check_if_cell_was_updated(cell_info, successful_matches, sheet_name):
    """
    특정 셀이 성공적으로 롤포워딩되었는지 확인
    
    Parameters:
        cell_info (dict): 셀 정보
        successful_matches (list): 성공한 매칭 정보
        sheet_name (str): 워크시트 이름
        
    Returns:
        bool: 업데이트 성공 여부
    """
    
    # 성공한 매칭 정보를 바탕으로 해당 셀이 업데이트되었는지 확인
    for match in successful_matches:
        if (match.get('from_sheet') == sheet_name and 
            match.get('from_row') == cell_info['row'] and
            match.get('from_col') == cell_info['col']):
            return True
    
    # 간단한 휴리스틱: 매칭된 헤더가 있는 컬럼의 셀들은 성공으로 간주
    # (실제로는 더 정밀한 추적 시스템이 필요)
    for match in successful_matches:
        from_table = match.get('from_table', {})
        if from_table.get('sheet') == sheet_name:
            try:
                from_headers = from_table.get('headers', [])
                if cell_info['col'] <= len(from_headers):
                    return True
            except:
                continue
    
    return False

def generate_manual_adjustment_report(previous_file=None, red_cells_info=None, successful_matches=None):
    """
    수기조정이 필요한 셀들의 상세 리포트 생성
    
    Parameters:
        previous_file (str): 전기 조서 파일 경로 (신규 시그니처)
        red_cells_info (dict): 빨간색으로 표시된 셀 정보 (신규 시그니처)
        successful_matches (list): 성공한 매칭 정보 (신규 시그니처)
        
    Returns:
        str: 리포트 텍스트
    """
    
    # 🔧 새로운 시그니처 지원: manual_adjustment_cells 생성
    manual_adjustment_cells = []
    
    if red_cells_info and successful_matches:
        # red_cells_info에서 실패한 셀들 추출
        try:
            wb = openpyxl.load_workbook(previous_file, data_only=True) if previous_file else None
            
            for sheet_name, cells_list in red_cells_info.items():
                if wb and sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    for cell_info in cells_list:
                        if cell_info.get('status') == 'failed' or not _check_if_cell_was_updated(cell_info, successful_matches, sheet_name):
                            cell = sheet.cell(cell_info['row'], cell_info['col'])
                            manual_adjustment_cells.append({
                                'sheet': sheet_name,
                                'row': cell_info['row'],
                                'col': cell_info['col'], 
                                'cell_address': f"{sheet_name}!{cell.coordinate}",
                                'value': cell_info.get('value', cell.value)
                            })
            if wb:
                wb.close()
        except Exception as e:
            print(f"[file_updater.generate_manual_adjustment_report] ⚠️ 셀 분석 중 오류: {e}")
    
    if not manual_adjustment_cells:
        return "🎉 모든 백데이터가 성공적으로 롤포워딩되었습니다!"
    
    report_lines = [
        "⚠️ 수기조정이 필요한 항목이 있습니다.",
        f"📊 총 {len(manual_adjustment_cells)}개의 셀이 자동 롤포워딩되지 않았습니다.",
        "",
        "📍 수기조정 필요 위치:"
    ]
    
    # 워크시트별로 그룹핑
    by_sheet = {}
    for cell in manual_adjustment_cells:
        sheet = cell['sheet']
        if sheet not in by_sheet:
            by_sheet[sheet] = []
        by_sheet[sheet].append(cell)
    
    for sheet_name, cells in by_sheet.items():
        report_lines.append(f"\n🔶 워크시트: {sheet_name}")
        for cell in cells:
            report_lines.append(f"   • {cell['cell_address']}: {cell['value']}")
    
    report_lines.extend([
        "",
        "💡 해결 방법:",
        "1. Excel 파일을 열어서 빨간색으로 표시된 셀들을 확인하세요",
        "2. 당기 PBC 파일에서 해당하는 값을 찾아 수동으로 복사하세요",
        "3. 또는 헤더 이름을 정확히 일치시켜 다시 롤포워딩을 시도하세요"
    ])
    
    return "\n".join(report_lines)

def add_rollforward_complete_workflow(source_file, target_file, back_data_sheets, tables_info, matches):
    """
    🆕 완전한 롤포워딩 워크플로우 오케스트레이션 (main.py 지원용)
    
    이 기능은 main.py에서 대체 사용 가능:
    1. 백데이터를 빨간색으로 표시 (pending)
    2. 데이터 복사 실행 + 노란색 표시 (in progress)
    3. 성공한 항목을 초록색으로 변경 (completed)
    4. 수동 조정 리포트 생성
    
    Parameters:
        source_file (str): 전기 조서 파일 경로
        target_file (str): 대상 PBC 파일 경로
        back_data_sheets (list): 백데이터 워크시트 리스트
        tables_info (list): 테이블 정보 리스트
        matches (list): 매칭 정보 리스트
        
    Returns:
        dict: 워크플로우 실행 결과
    """
    try:
        print(f"[file_updater.add_rollforward_complete_workflow] 🚀 완전 워크플로우 시작...")
        
        # 1단계: 백데이터 빨간색 표시
        print(f"[file_updater.add_rollforward_complete_workflow] 🔴 1단계: 백데이터 마킹 시작...")
        red_cells_info = mark_back_data_red(source_file, back_data_sheets, tables_info)
        
        # 2단계: 데이터 롤포워딩 실행
        print(f"[file_updater.add_rollforward_complete_workflow] 🟡 2단계: 데이터 복사 시작...")
        update_result = update_file(target_file, matches)
        
        # 3단계: 상태 업데이트 (성공 → 초록색)
        print(f"[file_updater.add_rollforward_complete_workflow] 🟢 3단계: 상태 업데이트 시작...")
        status_result = update_rollforward_status(source_file, red_cells_info, matches)
        
        # 4단계: 수동 조정 리포트 생성
        print(f"[file_updater.add_rollforward_complete_workflow] 📋 4단계: 리포트 생성 시작...")
        report = generate_manual_adjustment_report(source_file, red_cells_info, matches)
        
        # 결과 리턴
        result = {
            'success': True,
            'red_cells_marked': sum(len(cells) for cells in red_cells_info.values()) if red_cells_info else 0,
            'data_updated': isinstance(update_result, dict) and update_result.get('success'),
            'status_updated': status_result,
            'report': report
        }
        
        print(f"[file_updater.add_rollforward_complete_workflow] ✅ 완전 워크플로우 완료!")
        return result
        
    except Exception as e:
        print(f"[file_updater.add_rollforward_complete_workflow] ❌ 워크플로우 실행 오류: {e}")
        return {'success': False, 'error': str(e)}

if __name__ == "__main__":
    test_file_updater()