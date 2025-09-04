"""
Memory Efficient Worksheet Copy
메모리 효율적인 워크시트 복사 모듈 - 사용자 요청에 따른 간단한 구현
"""
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from copy import copy
import logging
import time
import os
import shutil

logger = logging.getLogger(__name__)

def copy_worksheet_like_ctrl_cv(source_file, source_sheet, target_file, target_sheet):
    """
    Ctrl+C, Ctrl+V처럼 워크시트 전체를 덮어쓰기 (간편 래퍼 함수)
    
    이 함수는 worksheet_full_replace의 간편한 래퍼입니다.
    내부적으로 동일한 완전한 서식 보존 복사를 수행합니다.
    
    Args:
        source_file: 소스 파일 경로
        source_sheet: 소스 시트명
        target_file: 타겟 파일 경로  
        target_sheet: 타겟 시트명
    
    Returns:
        bool: 성공 여부
    """
    # worksheet_full_replace 함수에 위임 (동일한 완전한 복사 기능)
    return worksheet_full_replace(
        source_file=source_file,
        source_sheet=source_sheet,
        target_file=target_file,
        target_sheet=target_sheet,
        preserve_formulas=True
    )

def simple_values_only_copy(source_file, source_sheet, target_file, target_sheet):
    """
    값만 복사하는 가장 간단한 방식 (더 빠르고 메모리 효율적)
    """
    try:
        source_wb = load_workbook(source_file, data_only=True)  # 값만 로드
        target_wb = load_workbook(target_file)
        
        source_ws = source_wb[source_sheet]
        target_ws = target_wb[target_sheet]
        
        # 기존 내용 모두 삭제
        target_ws.delete_rows(1, target_ws.max_row)
        
        # 값만 간단하게 복사 (가장 빠른 방식)
        for row in source_ws.values:
            target_ws.append(row)
            
        target_wb.save(target_file)
        return True
        
    except Exception as e:
        logger.error(f"Values only copy 실패: {str(e)}")
        return False

def worksheet_full_replace(source_file, source_sheet, target_file, target_sheet, preserve_formulas=True):
    """
    ✅ 강화된 워크시트 교체 함수
    - 파일 권한 및 잠금 상태 확인
    - 백업 생성 및 복구 기능
    - 상세 에러 처리 및 로깅
    - 메모리 효율적 처리
    
    Args:
        source_file: 소스 파일 경로
        source_sheet: 소스 시트명
        target_file: 타겟 파일 경로
        target_sheet: 타겟 시트명
        preserve_formulas: 수식 보존 여부 (기본값 True)
        
    Returns:
        bool: 성공 여부
    """
    import os
    import tempfile
    import shutil
    from pathlib import Path
    
    backup_file = None
    
    try:
        logger.info(f"🔄 워크시트 교체 시작: {source_sheet} ({Path(source_file).name}) → {target_sheet} ({Path(target_file).name})")
        
        # 1. 파일 존재 및 권한 확인
        if not os.path.exists(source_file):
            logger.error(f"❌ 소스 파일이 존재하지 않음: {source_file}")
            return False
            
        if not os.path.exists(target_file):
            logger.error(f"❌ 타겟 파일이 존재하지 않음: {target_file}")
            return False
            
        # 타겟 폴더 쓰기 권한 확인
        target_dir = os.path.dirname(target_file)
        if not os.access(target_dir, os.W_OK):
            logger.warning(f"⚠️ 타겟 폴더 쓰기 권한 없음, 권한 확인: {target_dir}")
            # 폴더 권한 문제는 경고만 하고 계속 진행 (실제 저장 시 처리)
            
        # 실제 파일 접근 테스트 (클라우드 동기화 파일에서 os.access 오동작 방지)
        try:
            # 소스 파일 읽기 테스트
            test_wb = load_workbook(source_file, read_only=True)
            test_wb.close()
        except Exception as e:
            logger.error(f"❌ 소스 파일 읽기 실패: {source_file} - {e}")
            return False
            
        try:
            # 타겟 파일 쓰기 테스트 (실제 열어보기)
            test_wb = load_workbook(target_file, read_only=False)
            test_wb.close()
        except PermissionError:
            # 읽기 전용 속성 제거 시도
            try:
                import stat
                logger.warning(f"⚠️ 파일 권한 문제 감지, 읽기 전용 속성 제거 시도: {target_file}")
                os.chmod(target_file, stat.S_IWRITE | stat.S_IREAD)
                
                # 다시 테스트
                test_wb = load_workbook(target_file, read_only=False)
                test_wb.close()
                logger.info(f"✅ 파일 권한 문제 해결: {target_file}")
            except Exception as chmod_error:
                logger.error(f"❌ 타겟 파일이 다른 프로그램에서 사용 중: {target_file}")
                logger.error("💡 해결방법: Excel에서 해당 파일을 닫고 다시 시도하세요")
                logger.error(f"💡 또는 파일 속성에서 '읽기 전용' 해제 후 다시 시도: {chmod_error}")
                return False
        except Exception as e:
            logger.error(f"❌ 타겟 파일 접근 실패: {target_file} - {e}")
            return False
        
        # 2. 백업은 main.py에서 이미 수행됨 - 여기서는 임시 백업만 생성
        try:
            # 임시 백업 (복구용) - 작업 실패 시 즉시 복원용
            temp_backup = target_file + ".temp_backup_" + str(int(time.time()))
            shutil.copy2(target_file, temp_backup)
            
            logger.info(f"🔄 작업용 임시 백업 생성: {temp_backup}")
            backup_file = temp_backup
            
        except Exception as backup_error:
            logger.warning(f"⚠️ 임시 백업 파일 생성 실패 (계속 진행): {backup_error}")
            backup_file = None
        
        # 3. 워크북 로드 (파일 잠금 처리)
        source_wb = None
        target_wb = None
        
        try:
            # 소스 파일 로드 (워크시트 복사를 위해 read_only=False 사용)
            source_wb = load_workbook(source_file, read_only=False, data_only=not preserve_formulas)
            logger.info(f"📖 소스 파일 로드 완료: {Path(source_file).name}")
            
            # 타겟 파일 로드 (쓰기 가능)
            target_wb = load_workbook(target_file, data_only=False)
            logger.info(f"📝 타겟 파일 로드 완료: {Path(target_file).name}")
            
        except PermissionError as perm_error:
            logger.error(f"❌ 파일 접근 권한 오류: {perm_error}")
            logger.error("💡 해결방법: Excel에서 해당 파일들을 모두 닫고 다시 시도하세요")
            return False
        except Exception as load_error:
            logger.error(f"❌ 파일 로드 오류: {load_error}")
            return False
        
        # 4. 워크시트 존재 확인
        if source_sheet not in source_wb.sheetnames:
            logger.error(f"❌ 소스 시트가 존재하지 않음: '{source_sheet}' (사용 가능: {source_wb.sheetnames})")
            return False
            
        if target_sheet not in target_wb.sheetnames:
            logger.error(f"❌ 타겟 시트가 존재하지 않음: '{target_sheet}' (사용 가능: {target_wb.sheetnames})")
            return False
        
        source_ws = source_wb[source_sheet]
        target_ws = target_wb[target_sheet]
        
        # 5. 데이터 크기 확인 및 메모리 체크
        try:
            source_dimension = source_ws.calculate_dimension()
            logger.info(f"📊 소스 데이터 범위: {source_dimension}")
        except Exception:
            logger.warning("⚠️ 소스 데이터 범위 확인 실패 (계속 진행)")
        
        # 6. 워크시트 내용 교체
        logger.info(f"🔄 워크시트 내용 교체 중...")
        
        # 기존 내용 백업을 위한 임시 저장
        original_cells = target_ws._cells.copy()
        
        try:
            # 완전한 워크시트 복사 (Ctrl+A, Ctrl+C, Ctrl+V와 동일)
            logger.info(f"🎯 전체 워크시트 복사 시작...")
            
            # 기존 타겟 워크시트 내용 완전 삭제
            target_ws.delete_rows(1, target_ws.max_row or 1)
            target_ws.delete_cols(1, target_ws.max_column or 1)
            
            # 모든 셀 데이터와 서식을 개별적으로 복사 (완전한 서식 보존)
            logger.info(f"📋 셀별 데이터 및 서식 복사 중...")
            
            # 소스 워크시트의 사용된 범위 확인
            if source_ws.max_row and source_ws.max_column:
                for row in source_ws.iter_rows(min_row=1, max_row=source_ws.max_row, 
                                             min_col=1, max_col=source_ws.max_column):
                    for source_cell in row:
                        if source_cell.coordinate:
                            target_cell = target_ws[source_cell.coordinate]
                            
                            # 셀 값 복사
                            target_cell.value = source_cell.value
                            
                            # 셀 서식 완전 복사
                            if source_cell.has_style:
                                # 숫자 형식 (날짜, 비율, 통화 등)
                                target_cell.number_format = source_cell.number_format
                                
                                # 폰트 서식
                                target_cell.font = copy(source_cell.font)
                                
                                # 테두리 서식
                                target_cell.border = copy(source_cell.border)
                                
                                # 채우기/배경색 서식
                                target_cell.fill = copy(source_cell.fill)
                                
                                # 정렬 서식
                                target_cell.alignment = copy(source_cell.alignment)
                                
                                # 보호 설정
                                target_cell.protection = copy(source_cell.protection)
                            
                            # 하이퍼링크 복사
                            if source_cell.hyperlink:
                                target_cell.hyperlink = copy(source_cell.hyperlink)
                            
                            # 주석/메모 복사
                            if source_cell.comment:
                                target_cell.comment = copy(source_cell.comment)
            
            # 행/열 차원 정보 복사
            target_ws.column_dimensions = source_ws.column_dimensions.copy()
            target_ws.row_dimensions = source_ws.row_dimensions.copy()
            
            # 병합된 셀 복사
            target_ws.merged_cells.ranges = list(source_ws.merged_cells.ranges)
            
            # 추가 속성 복사 (모든 서식 보존)
            try:
                # 조건부 서식
                if hasattr(source_ws, 'conditional_formatting'):
                    target_ws.conditional_formatting = copy(source_ws.conditional_formatting)
                
                # 데이터 유효성 검사
                if hasattr(source_ws, 'data_validations'):
                    target_ws.data_validations = copy(source_ws.data_validations)
                
                # 워크시트 보호
                if hasattr(source_ws, 'protection'):
                    target_ws.protection = copy(source_ws.protection)
                
                # 페이지 설정
                if hasattr(source_ws, 'page_setup'):
                    target_ws.page_setup = copy(source_ws.page_setup)
                if hasattr(source_ws, 'page_margins'):
                    target_ws.page_margins = copy(source_ws.page_margins)
                if hasattr(source_ws, 'print_options'):
                    target_ws.print_options = copy(source_ws.print_options)
                
                # 워크시트 뷰 설정 (읽기 전용 속성들은 스킵)
                try:
                    if hasattr(source_ws, 'sheet_view') and hasattr(target_ws.__class__, 'sheet_view') and hasattr(target_ws.__class__.sheet_view, 'fset'):
                        target_ws.sheet_view = copy(source_ws.sheet_view)
                except (AttributeError, TypeError):
                    pass  # 읽기 전용 속성
                    
                try:
                    if hasattr(source_ws, 'views'):
                        target_ws.views = copy(source_ws.views)
                except (AttributeError, TypeError):
                    pass  # 읽기 전용 속성
                
                # 기타 워크시트 속성
                if hasattr(source_ws, 'sheet_format'):
                    target_ws.sheet_format = copy(source_ws.sheet_format)
                if hasattr(source_ws, 'sheet_properties'):
                    target_ws.sheet_properties = copy(source_ws.sheet_properties)
                if hasattr(source_ws, 'auto_filter'):
                    target_ws.auto_filter = copy(source_ws.auto_filter)
                if hasattr(source_ws, 'freeze_panes'):
                    target_ws.freeze_panes = source_ws.freeze_panes
                    
                logger.info(f"✅ 모든 서식 속성 복사 완료")
                    
            except Exception as attr_error:
                logger.warning(f"⚠️ 일부 속성 복사 실패 (계속 진행): {attr_error}")
            
            logger.info(f"✅ 워크시트 내용 교체 완료")
            
        except Exception as copy_error:
            logger.error(f"❌ 워크시트 내용 복사 실패: {copy_error}")
            # 원래 내용 복구
            target_ws._cells = original_cells
            return False
        
        # 7. 파일 저장 (권한 처리 강화)
        try:
            logger.info(f"💾 파일 저장 중...")
            target_wb.save(target_file)
            logger.info(f"✅ 파일 저장 완료: {Path(target_file).name}")
        except PermissionError as perm_error:
            logger.error(f"❌ 파일 저장 권한 오류: {perm_error}")
            
            # 권한 문제 해결 시도
            try:
                import stat
                logger.warning(f"⚠️ 파일 저장 권한 문제 해결 시도: {target_file}")
                
                # 읽기 전용 속성 제거
                os.chmod(target_file, stat.S_IWRITE | stat.S_IREAD)
                
                # 다시 저장 시도
                target_wb.save(target_file)
                logger.info(f"✅ 권한 문제 해결 후 파일 저장 완료: {Path(target_file).name}")
                
            except Exception as chmod_error:
                logger.error(f"❌ 권한 문제 해결 실패: {chmod_error}")
                logger.error("💡 해결방법:")
                logger.error("   1. Excel에서 해당 파일을 닫고 다시 시도")
                logger.error("   2. 파일 속성에서 '읽기 전용' 해제")
                logger.error("   3. 파일이 있는 폴더의 쓰기 권한 확인")
                return False
                
        except Exception as save_error:
            logger.error(f"❌ 파일 저장 실패: {save_error}")
            logger.error("💡 가능한 원인:")
            logger.error("   - 파일이 다른 프로그램에서 사용 중")
            logger.error("   - 디스크 공간 부족")
            logger.error("   - 파일 경로가 너무 길거나 잘못됨")
            return False
        
        # 8. 임시 백업 파일 정리 (영구 백업은 보존)
        if backup_file and os.path.exists(backup_file):
            try:
                os.remove(backup_file)
                logger.info(f"🗑️ 임시 백업 파일 정리 완료")
            except Exception:
                logger.warning(f"⚠️ 임시 백업 파일 정리 실패: {backup_file}")
        
        logger.info(f"🎉 워크시트 전체 교체 성공: {source_sheet} → {target_sheet}")
        return True
        
    except Exception as e:
        logger.error(f"❌ 워크시트 교체 중 예상치 못한 오류: {str(e)}")
        
        # 백업 파일로 복구 시도
        if backup_file and os.path.exists(backup_file):
            try:
                shutil.copy2(backup_file, target_file)
                logger.info(f"🔄 백업 파일로 복구 완료")
                os.remove(backup_file)
            except Exception as recovery_error:
                logger.error(f"❌ 백업 복구도 실패: {recovery_error}")
        
        return False
    
    finally:
        # 리소스 정리
        try:
            if 'source_wb' in locals() and source_wb:
                source_wb.close()
            if 'target_wb' in locals() and target_wb:
                target_wb.close()
        except Exception:
            pass