# -*- coding: utf-8 -*-
"""
롤포워딩 로그 워크시트 생성 모듈
"""
import os
from datetime import datetime
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

def create_rollforward_log_worksheet(target_file, log_data):
    """
    롤포워딩 완료 후 로그 워크시트를 대상 조서에 생성
    
    Args:
        target_file (str): 롤포워딩 대상 조서 파일 경로
        log_data (dict): 롤포워딩 처리 결과 데이터
        
    Returns:
        bool: 로그 워크시트 생성 성공 여부
    """
    try:
        logger.info(f"📊 롤포워딩 로그 워크시트 생성 시작: {Path(target_file).name}")
        
        # 파일 존재 확인
        if not os.path.exists(target_file):
            logger.error(f"❌ 대상 파일이 존재하지 않음: {target_file}")
            return False
        
        # 파일 권한 확인 및 읽기 전용 속성 제거
        try:
            # 파일에 실제로 접근해보기 (os.access보다 정확)
            with open(target_file, 'r+b') as test_file:
                pass
        except PermissionError:
            # 읽기 전용 속성 제거 시도
            try:
                import stat
                current_mode = os.stat(target_file).st_mode
                os.chmod(target_file, current_mode | stat.S_IWRITE)
                logger.info(f"📝 읽기 전용 속성 제거: {Path(target_file).name}")
            except Exception as chmod_error:
                logger.error(f"❌ 파일 권한 수정 실패: {chmod_error}")
                logger.error("💡 해결방법: Excel에서 해당 파일을 닫고 파일 속성에서 읽기 전용을 해제하세요")
                return False
        
        # 워크북 열기
        try:
            wb = load_workbook(target_file)
        except PermissionError as perm_error:
            logger.error(f"❌ 파일 접근 권한 오류: {perm_error}")
            logger.error("💡 해결방법: Excel에서 해당 파일을 모두 닫고 다시 시도하세요")
            return False
        
        # 로그 워크시트명 (중복 방지, 31자 제한)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_sheet_name = f"RF_Log_{timestamp}"
        
        # 기존 로그 시트가 있으면 제거 (최신 로그만 유지)
        existing_logs = [ws.title for ws in wb.worksheets if ws.title.startswith("RF_Log")]
        for old_log in existing_logs:
            if old_log in wb.sheetnames:
                wb.remove(wb[old_log])
                logger.info(f"🗑️ 기존 로그 워크시트 제거: {old_log}")
        
        # 새 로그 워크시트 생성
        log_ws = wb.create_sheet(title=log_sheet_name)
        
        # 헤더 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # 1. 롤포워딩 요약 정보
        _create_summary_section(log_ws, log_data, header_font, header_fill, border, center_align)
        
        # 2. 백데이터 워크시트 처리 결과
        _create_backdata_section(log_ws, log_data, header_font, header_fill, border, center_align)
        
        # 3. 테이블 단위 처리 결과
        _create_table_section(log_ws, log_data, header_font, header_fill, border, center_align)
        
        # 4. 성공/실패 상세 내역
        _create_details_section(log_ws, log_data, header_font, header_fill, border, center_align)
        
        # 열 너비 자동 조정
        _auto_adjust_column_width(log_ws)
        
        # 워크북 저장
        wb.save(target_file)
        wb.close()
        
        logger.info(f"✅ 롤포워딩 로그 워크시트 생성 완료: {log_sheet_name}")
        return True
        
    except Exception as e:
        logger.error(f"❌ 롤포워딩 로그 워크시트 생성 실패: {e}")
        return False

def _create_summary_section(ws, log_data, header_font, header_fill, border, center_align):
    """롤포워딩 요약 정보 섹션 생성"""
    current_row = 1
    
    # 제목
    ws.cell(row=current_row, column=1, value="Roll-Forwarding 처리 요약")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2
    
    # 요약 정보
    summary_data = [
        ["처리 시간", log_data.get('timestamp', datetime.now().strftime("%Y-%m-%d %H:%M:%S"))],
        ["대상 조서", log_data.get('target_file_name', 'N/A')],
        ["당기 PBC 폴더", log_data.get('source_folder', 'N/A')],
        ["총 처리 워크시트", str(log_data.get('total_worksheets', 0)) + "개"],
        ["성공한 워크시트", str(log_data.get('success_worksheets', 0)) + "개"],
        ["실패한 워크시트", str(log_data.get('failed_worksheets', 0)) + "개"],
    ]
    
    for row_data in summary_data:
        ws.cell(row=current_row, column=1, value=row_data[0])
        ws.cell(row=current_row, column=2, value=row_data[1])
        
        # 스타일 적용
        ws.cell(row=current_row, column=1).font = Font(bold=True)
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        for col in range(1, 3):
            ws.cell(row=current_row, column=col).border = border
            
        current_row += 1
    
    return current_row + 1

def _create_backdata_section(ws, log_data, header_font, header_fill, border, center_align):
    """백데이터 워크시트 처리 섹션 생성"""
    current_row = ws.max_row + 2
    
    # 섹션 제목
    ws.cell(row=current_row, column=1, value="백데이터 워크시트 처리 결과 (워크시트 전체 복사)")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2
    
    # 헤더
    headers = ["순번", "대상 워크시트", "소스 파일", "소스 워크시트", "매칭 신뢰도", "처리 결과"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    current_row += 1
    
    # 백데이터 처리 결과
    backdata_results = log_data.get('backdata_results', [])
    for idx, result in enumerate(backdata_results, 1):
        row_data = [
            idx,
            result.get('target_worksheet', 'N/A'),
            result.get('source_file', 'N/A'),
            result.get('source_worksheet', 'N/A'),
            f"{result.get('confidence', 0):.1%}",
            "성공" if result.get('success', False) else "실패"
        ]
        
        for col, data in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=data)
            cell.border = border
            
            # 성공/실패 색상 적용
            if col == 6:  # 처리 결과 컬럼
                if "성공" in str(data):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        current_row += 1
    
    return current_row

def _create_table_section(ws, log_data, header_font, header_fill, border, center_align):
    """테이블 단위 처리 섹션 생성"""
    current_row = ws.max_row + 2
    
    # 섹션 제목
    ws.cell(row=current_row, column=1, value="테이블 단위 처리 결과")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:G{current_row}')
    current_row += 2
    
    # 헤더
    headers = ["순번", "대상 워크시트", "테이블 범위", "소스 파일", "소스 테이블", "매칭률", "처리 결과"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align
    
    current_row += 1
    
    # 테이블 처리 결과
    table_results = log_data.get('table_results', [])
    for idx, result in enumerate(table_results, 1):
        row_data = [
            idx,
            result.get('target_worksheet', 'N/A'),
            result.get('table_range', 'N/A'),
            result.get('source_file', 'N/A'),
            result.get('source_table', 'N/A'),
            f"{result.get('match_rate', 0):.1%}",
            "성공" if result.get('success', False) else "실패"
        ]
        
        for col, data in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=data)
            cell.border = border
            
            # 성공/실패 색상 적용
            if col == 7:  # 처리 결과 컬럼
                if "성공" in str(data):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        current_row += 1
    
    return current_row

def _create_details_section(ws, log_data, header_font, header_fill, border, center_align):
    """처리 상세 내역 섹션 생성"""
    current_row = ws.max_row + 2
    
    # 섹션 제목
    ws.cell(row=current_row, column=1, value="처리 상세 내역")
    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{current_row}:D{current_row}')
    current_row += 2
    
    # 실패 사유 및 해결 방법
    failures = log_data.get('failures', [])
    if failures:
        ws.cell(row=current_row, column=1, value="실패 사유 및 해결 방법:")
        ws.cell(row=current_row, column=1).font = Font(bold=True, color="FF0000")
        current_row += 1
        
        for failure in failures:
            ws.cell(row=current_row, column=1, value=f"• {failure.get('reason', 'N/A')}")
            ws.cell(row=current_row, column=2, value=failure.get('solution', 'N/A'))
            current_row += 1
    
    # 처리 통계
    current_row += 1
    stats = log_data.get('statistics', {})
    ws.cell(row=current_row, column=1, value="처리 통계:")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    current_row += 1
    
    stats_data = [
        f"• 총 처리 시간: {stats.get('total_time', 'N/A')}",
        f"• 복사된 데이터 셀 수: {stats.get('copied_cells', 0):,}개",
        f"• 생성된 백업 파일: {stats.get('backup_files', 0)}개",
        f"• 처리된 테이블 수: {stats.get('processed_tables', 0)}개"
    ]
    
    for stat in stats_data:
        ws.cell(row=current_row, column=1, value=stat)
        current_row += 1

def _auto_adjust_column_width(ws):
    """열 너비 자동 조정"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                # MergedCell인지 확인하고 실제 셀만 처리
                if hasattr(cell, 'value') and cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # 최대 50자로 제한
        ws.column_dimensions[column_letter].width = adjusted_width