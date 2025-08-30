"""
직관적 승인 인터페이스
워크시트 감지 결과를 기반으로 한 사용자 확인 UI

Features:
- 감지된 워크시트 목록 (테이블 형태)
- 신뢰도 기반 컬러 코딩 (높음=초록, 중간=노랑, 낮음=빨강)
- 상호작용 요소들 (체크박스, 미리보기, 설정 저장)
- 접근성 고려 (키보드 네비게이션, 명확한 라벨링)
- 멘토링 요소 (교육적 설명)
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import os
from pathlib import Path
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass

# Local imports
try:
    from .config_manager import ConfigManager
except ImportError:
    # Fallback for direct execution
    from config_manager import ConfigManager

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl이 없습니다. 미리보기 기능이 제한됩니다.")


@dataclass
class DetectionResult:
    """간단한 감지 결과 데이터 클래스"""
    worksheet_name: str
    confidence_score: float
    detection_method: str
    reasoning: str
    is_backdata: bool
    processing_time: float


@dataclass
class WorksheetChoice:
    """사용자 워크시트 선택 정보"""
    worksheet_name: str
    is_selected: bool
    detection_result: DetectionResult
    user_override: bool = False  # 사용자가 감지 결과를 변경했는지


class WorksheetConfirmationDialog:
    """
    직관적 승인 인터페이스 - 메인 클래스
    
    Features:
    - 워크시트 감지 결과 표시
    - 사용자 확인 및 수정
    - 접근성 지원
    - 교육적 가이드
    """
    
    def __init__(self, excel_path: Path, parent=None):
        self.excel_path = Path(excel_path)
        self.parent = parent
        
        # Core components
        self.config_manager = ConfigManager()
        
        # UI state
        self.worksheet_choices: List[WorksheetChoice] = []
        self.selected_count = 0
        self.dialog_result = None
        
        # UI components
        self.root = None
        self.main_frame = None
        self.worksheet_vars = []  # Checkbutton variables
        self.worksheet_widgets = []  # UI widgets for each worksheet
        
        # Accessibility features
        self.current_focus_index = 0
        self.keyboard_navigation = True
    
    def show_confirmation_dialog(self) -> Tuple[List[str], List[str]]:
        """
        메인 확인 대화상자 표시
        
        Returns:
            Tuple[List[str], List[str]]: (본조서 워크시트들, 백데이터 워크시트들)
        """
        try:
            # 1. 워크시트 감지 실행
            self._detect_worksheets()
            
            # 2. UI 생성 및 표시
            self._create_dialog()
            
            # 3. 사용자 응답 대기
            self.root.wait_window()
            
            # 4. 결과 반환
            if self.dialog_result:
                return self._process_results()
            else:
                return [], []  # 사용자가 취소함
                
        except Exception as e:
            messagebox.showerror("오류", f"확인 대화상자 생성 중 오류가 발생했습니다:\n{e}")
            return [], []
    
    def _detect_worksheets(self):
        """워크시트 수동 선택 (간단한 MVP 버전)"""
        try:
            # 간단하게 모든 워크시트를 수동 선택으로 설정
            self._manual_worksheet_detection()
                
        except Exception as e:
            # 오류 발생 시에도 수동 선택으로 fallback
            messagebox.showwarning("오류", 
                f"워크시트 목록을 가져오는 중 오류가 발생했습니다.\n오류: {e}")
            self.worksheet_choices = []
    
    def _manual_worksheet_detection(self):
        """수동 워크시트 감지 (fallback)"""
        try:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl이 필요합니다.")
                
            workbook = openpyxl.load_workbook(self.excel_path, read_only=True)
            
            self.worksheet_choices = []
            for ws_name in workbook.sheetnames:
                # 기본 DetectionResult 생성
                result = DetectionResult(
                    worksheet_name=ws_name,
                    confidence_score=0.3,
                    detection_method="Manual",
                    reasoning="수동 선택 모드",
                    is_backdata=False,
                    processing_time=0.0
                )
                
                choice = WorksheetChoice(
                    worksheet_name=ws_name,
                    is_selected=False,
                    detection_result=result
                )
                self.worksheet_choices.append(choice)
                
            workbook.close()
            
        except Exception as e:
            messagebox.showerror("오류", f"워크시트 목록을 가져올 수 없습니다:\n{e}")
            self.worksheet_choices = []
    
    def _create_dialog(self):
        """메인 대화상자 UI 생성"""
        # Root window 설정
        self.root = tk.Toplevel(self.parent) if self.parent else tk.Tk()
        self.root.title("본 조서 선택 시스템")
        self.root.geometry("900x700")
        self.root.resizable(True, True)
        
        # 창을 화면 중앙에 위치
        self.root.transient(self.parent)
        self.root.grab_set()
        
        # 메인 프레임
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky="nsew")
        
        # Grid 설정
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        self.main_frame.columnconfigure(1, weight=1)
        
        # UI 컴포넌트들 생성
        self._create_header()
        self._create_educational_guide()
        self._create_worksheet_table()
        self._create_control_buttons()
        
        # 접근성 설정
        self._setup_accessibility()
        
        # 초기 포커스 설정
        self.root.focus_set()
    
    def _create_header(self):
        """헤더 영역 생성"""
        header_frame = ttk.Frame(self.main_frame)
        header_frame.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        
        # 제목
        title_label = ttk.Label(header_frame, 
            text="본 조서 워크시트 선택",
            font=("맑은 고딕", 16, "bold"))
        title_label.grid(row=0, column=0, sticky="w")
        
        # 파일 경로 표시
        file_label = ttk.Label(header_frame,
            text=f"파일: {self.excel_path.name}",
            font=("맑은 고딕", 10))
        file_label.grid(row=1, column=0, sticky="w", pady=(5, 0))
        
        # 감지 통계
        total_worksheets = len(self.worksheet_choices)
        auto_detected = sum(1 for choice in self.worksheet_choices if choice.detection_result.confidence_score > 0.7)
        
        stats_label = ttk.Label(header_frame,
            text=f"총 워크시트: {total_worksheets}개 | 자동 감지: {auto_detected}개",
            font=("맑은 고딕", 10))
        stats_label.grid(row=2, column=0, sticky="w", pady=(2, 0))
    
    def _create_educational_guide(self):
        """교육적 가이드 영역 생성"""
        guide_frame = ttk.LabelFrame(self.main_frame, text="💡 가이드", padding="10")
        guide_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        
        guide_text = """
🎯 목적: 전기 조서에서 본 조서(최종 재무제표) 워크시트를 선택합니다.

📊 색상 의미:
  🟢 초록색 - 높은 신뢰도 (90%+): 자동 감지 결과를 신뢰할 수 있습니다
  🟡 노란색 - 중간 신뢰도 (50~90%): 검토 후 결정하시기 바랍니다  
  🔴 빨간색 - 낮은 신뢰도 (50% 미만): 수동으로 판단이 필요합니다

✅ 체크된 워크시트 = 본 조서 (최종 재무제표)
❌ 체크 안된 워크시트 = 백데이터 (이월 대상)

⌨️  키보드: Space(선택), Tab(이동), Enter(미리보기)"""
        
        guide_label = ttk.Label(guide_frame, text=guide_text, 
            font=("맑은 고딕", 9), justify="left")
        guide_label.grid(row=0, column=0, sticky="w")
    
    def _create_worksheet_table(self):
        """워크시트 테이블 생성"""
        table_frame = ttk.LabelFrame(self.main_frame, text="워크시트 목록", padding="5")
        table_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", pady=(0, 10))
        self.main_frame.rowconfigure(2, weight=1)
        
        # 전체 선택/해제 버튼
        control_frame = ttk.Frame(table_frame)
        control_frame.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        
        select_all_btn = ttk.Button(control_frame, text="전체 선택", 
            command=self._select_all_worksheets, width=12)
        select_all_btn.grid(row=0, column=0, padx=(0, 5))
        
        deselect_all_btn = ttk.Button(control_frame, text="전체 해제",
            command=self._deselect_all_worksheets, width=12) 
        deselect_all_btn.grid(row=0, column=1, padx=(0, 5))
        
        # 선택 카운터
        self.selection_label = ttk.Label(control_frame, text="")
        self.selection_label.grid(row=0, column=2, padx=(20, 0))
        
        # 스크롤 가능한 워크시트 리스트
        canvas_frame = ttk.Frame(table_frame)
        canvas_frame.grid(row=1, column=0, sticky="nsew")
        table_frame.rowconfigure(1, weight=1)
        table_frame.columnconfigure(0, weight=1)
        
        # Canvas와 Scrollbar
        canvas = tk.Canvas(canvas_frame, height=300)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Grid 배치
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas_frame.rowconfigure(0, weight=1)
        canvas_frame.columnconfigure(0, weight=1)
        
        # 워크시트 항목들 생성
        self._create_worksheet_items(scrollable_frame)
        
        # 마우스 휠 바인딩
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # 선택 카운터 업데이트
        self._update_selection_count()
    
    def _create_worksheet_items(self, parent_frame):
        """개별 워크시트 항목들 생성"""
        self.worksheet_vars = []
        self.worksheet_widgets = []
        
        # 헤더 행
        headers = ["선택", "워크시트명", "신뢰도", "추천", "처리방식", "미리보기"]
        for i, header in enumerate(headers):
            label = ttk.Label(parent_frame, text=header, font=("맑은 고딕", 9, "bold"))
            label.grid(row=0, column=i, padx=5, pady=5, sticky="w")
        
        # 구분선
        separator = ttk.Separator(parent_frame, orient="horizontal")
        separator.grid(row=1, column=0, columnspan=6, sticky="ew", padx=5, pady=2)
        
        # 각 워크시트 항목
        for idx, choice in enumerate(self.worksheet_choices, start=2):
            self._create_single_worksheet_row(parent_frame, idx, choice)
    
    def _create_single_worksheet_row(self, parent, row, choice: WorksheetChoice):
        """단일 워크시트 행 생성"""
        result = choice.detection_result
        
        # 1. 체크박스
        var = tk.BooleanVar(value=choice.is_selected)
        var.trace("w", self._on_selection_changed)
        
        checkbox = ttk.Checkbutton(parent, variable=var, 
            command=lambda: self._on_worksheet_toggled(choice, var))
        checkbox.grid(row=row, column=0, padx=5, pady=2)
        
        # 접근성: 라벨 설정
        checkbox.configure(text="", 
            command=lambda: self._on_worksheet_toggled(choice, var))
        
        # 2. 워크시트명 (색상 배경으로 신뢰도 표시)
        ws_frame = ttk.Frame(parent)
        ws_frame.grid(row=row, column=1, padx=5, pady=2, sticky="ew")
        
        # 신뢰도 기반 배경색 결정
        confidence = result.confidence_score
        if confidence >= 0.9:
            bg_color = "#d4edda"  # 연한 초록
            text_color = "#155724"
        elif confidence >= 0.5:
            bg_color = "#fff3cd"  # 연한 노랑  
            text_color = "#856404"
        else:
            bg_color = "#f8d7da"  # 연한 빨강
            text_color = "#721c24"
        
        ws_label = tk.Label(ws_frame, text=result.worksheet_name,
            bg=bg_color, fg=text_color, font=("맑은 고딕", 9, "bold"),
            padx=8, pady=4, relief="solid", borderwidth=1)
        ws_label.pack(fill="x")
        
        # 3. 신뢰도 점수
        confidence_text = f"{confidence:.1%}"
        confidence_label = ttk.Label(parent, text=confidence_text, 
            font=("맑은 고딕", 9))
        confidence_label.grid(row=row, column=2, padx=5, pady=2)
        
        # 4. AI 추천
        recommendation = "백데이터" if result.is_backdata else "본 조서"
        rec_label = ttk.Label(parent, text=recommendation,
            font=("맑은 고딕", 9, "bold"),
            foreground="#007bff" if result.is_backdata else "#6c757d")
        rec_label.grid(row=row, column=3, padx=5, pady=2)
        
        # 5. 처리 방식 (감지 방법)
        method = result.detection_method.replace("Level ", "L")
        method_label = ttk.Label(parent, text=method,
            font=("맑은 고딕", 8))
        method_label.grid(row=row, column=4, padx=5, pady=2)
        
        # 6. 미리보기 버튼
        preview_btn = ttk.Button(parent, text="미리보기", width=10,
            command=lambda ws=result.worksheet_name: self._show_worksheet_preview(ws))
        preview_btn.grid(row=row, column=5, padx=5, pady=2)
        
        # 위젯들 저장 (접근성을 위해)
        widget_group = {
            'checkbox': checkbox,
            'var': var,
            'choice': choice,
            'preview_btn': preview_btn,
            'row': row
        }
        self.worksheet_vars.append(var)
        self.worksheet_widgets.append(widget_group)
    
    def _create_control_buttons(self):
        """하단 컨트롤 버튼들 생성"""
        button_frame = ttk.Frame(self.main_frame)
        button_frame.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        
        # 왼쪽: 설정 저장
        left_frame = ttk.Frame(button_frame)
        left_frame.grid(row=0, column=0, sticky="w")
        
        save_config_btn = ttk.Button(left_frame, text="설정 저장", 
            command=self._save_to_whitelist, width=12)
        save_config_btn.grid(row=0, column=0, padx=(0, 10))
        
        # 설정 저장 설명
        save_info = ttk.Label(left_frame, 
            text="현재 선택을 화이트리스트에 저장하여 다음에 자동 적용",
            font=("맑은 고딕", 8), foreground="#6c757d")
        save_info.grid(row=0, column=1)
        
        # 오른쪽: 확인/취소
        right_frame = ttk.Frame(button_frame)
        right_frame.grid(row=0, column=1, sticky="e")
        
        cancel_btn = ttk.Button(right_frame, text="취소", 
            command=self._on_cancel, width=12)
        cancel_btn.grid(row=0, column=0, padx=(0, 10))
        
        confirm_btn = ttk.Button(right_frame, text="확인", 
            command=self._on_confirm, width=12)
        confirm_btn.grid(row=0, column=1)
        confirm_btn.focus_set()  # 기본 포커스
        
        # Grid 설정
        button_frame.columnconfigure(1, weight=1)
    
    def _setup_accessibility(self):
        """접근성 기능 설정"""
        # 키보드 바인딩
        self.root.bind("<KeyPress>", self._on_key_press)
        self.root.bind("<Tab>", self._on_tab_navigation)
        self.root.bind("<Shift-Tab>", self._on_shift_tab_navigation)
        
        # ESC로 취소
        self.root.bind("<Escape>", lambda e: self._on_cancel())
        
        # Enter로 확인
        self.root.bind("<Return>", lambda e: self._on_confirm())
        
        # F1으로 도움말
        self.root.bind("<F1>", self._show_help)
    
    # Event Handlers
    def _on_selection_changed(self, *args):
        """선택 변경 이벤트"""
        self._update_selection_count()
    
    def _on_worksheet_toggled(self, choice: WorksheetChoice, var: tk.BooleanVar):
        """워크시트 선택 토글"""
        choice.is_selected = var.get()
        
        # AI 추천과 다르면 사용자 오버라이드 표시
        if choice.is_selected != choice.detection_result.is_backdata:
            choice.user_override = True
        else:
            choice.user_override = False
        
        self._update_selection_count()
    
    def _update_selection_count(self):
        """선택 카운터 업데이트"""
        selected_count = sum(1 for choice in self.worksheet_choices if choice.is_selected)
        total_count = len(self.worksheet_choices)
        
        override_count = sum(1 for choice in self.worksheet_choices if choice.user_override)
        
        text = f"본 조서 선택: {selected_count}/{total_count}"
        if override_count > 0:
            text += f" (사용자 수정: {override_count}개)"
        
        self.selection_label.config(text=text)
    
    def _select_all_worksheets(self):
        """모든 워크시트 선택"""
        for var in self.worksheet_vars:
            var.set(True)
        for choice in self.worksheet_choices:
            choice.is_selected = True
        self._update_selection_count()
    
    def _deselect_all_worksheets(self):
        """모든 워크시트 선택 해제"""
        for var in self.worksheet_vars:
            var.set(False)
        for choice in self.worksheet_choices:
            choice.is_selected = False
        self._update_selection_count()
    
    def _show_worksheet_preview(self, worksheet_name: str):
        """워크시트 미리보기 표시"""
        if not OPENPYXL_AVAILABLE:
            messagebox.showinfo("미리보기", "openpyxl이 필요합니다.")
            return
        
        try:
            # 미리보기 창 생성
            preview_window = tk.Toplevel(self.root)
            preview_window.title(f"미리보기: {worksheet_name}")
            preview_window.geometry("800x500")
            preview_window.transient(self.root)
            
            # 로딩 표시
            loading_label = ttk.Label(preview_window, text="데이터를 읽는 중...", 
                font=("맑은 고딕", 12))
            loading_label.pack(expand=True)
            
            preview_window.update()
            
            # Excel 데이터 읽기 (첫 10행만)
            workbook = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            worksheet = workbook[worksheet_name]
            
            # 데이터 수집
            preview_data = []
            max_rows = min(worksheet.max_row or 10, 10)
            max_cols = min(worksheet.max_column or 10, 10)
            
            for row in range(1, max_rows + 1):
                row_data = []
                for col in range(1, max_cols + 1):
                    cell = worksheet.cell(row=row, column=col)
                    value = str(cell.value) if cell.value is not None else ""
                    # 긴 텍스트는 자르기
                    if len(value) > 30:
                        value = value[:27] + "..."
                    row_data.append(value)
                preview_data.append(row_data)
            
            workbook.close()
            
            # 로딩 라벨 제거
            loading_label.destroy()
            
            # TreeView로 테이블 표시
            tree_frame = ttk.Frame(preview_window)
            tree_frame.pack(fill="both", expand=True, padx=10, pady=10)
            
            # Treeview 생성
            columns = [f"열{i+1}" for i in range(max_cols)]
            tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=15)
            
            # 컬럼 헤더 설정
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, width=100, minwidth=80)
            
            # 데이터 삽입
            for i, row_data in enumerate(preview_data):
                tree.insert("", "end", iid=i, values=row_data)
            
            # 스크롤바
            v_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
            h_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=tree.xview)
            tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
            
            # Grid 배치
            tree.grid(row=0, column=0, sticky="nsew")
            v_scrollbar.grid(row=0, column=1, sticky="ns")
            h_scrollbar.grid(row=1, column=0, sticky="ew")
            
            tree_frame.rowconfigure(0, weight=1)
            tree_frame.columnconfigure(0, weight=1)
            
            # 닫기 버튼
            close_btn = ttk.Button(preview_window, text="닫기", 
                command=preview_window.destroy)
            close_btn.pack(pady=(0, 10))
            
        except Exception as e:
            messagebox.showerror("미리보기 오류", f"미리보기를 생성할 수 없습니다:\n{e}")
    
    def _save_to_whitelist(self):
        """현재 선택을 화이트리스트에 저장"""
        try:
            patterns_to_save = []
            
            for choice in self.worksheet_choices:
                if choice.user_override:  # 사용자가 수정한 항목만
                    pattern_config = {
                        "pattern": f"^{choice.worksheet_name}$",  # 정확한 매칭
                        "strategy": "worksheet_level" if choice.is_selected else "table_level",
                        "auto_approve": True,
                        "comment": f"사용자 설정: {'백데이터' if choice.is_selected else '본조서'}"
                    }
                    patterns_to_save.append(pattern_config)
            
            if patterns_to_save:
                # 기존 화이트리스트에 추가
                current_config = self.config_manager.load_whitelist_config()
                current_patterns = current_config.get("worksheet_patterns", [])
                current_patterns.extend(patterns_to_save)
                
                updated_config = {
                    "version": current_config.get("version", "1.0.0"),
                    "worksheet_patterns": current_patterns
                }
                
                self.config_manager.save_whitelist_config(updated_config)
                messagebox.showinfo("저장 완료", 
                    f"{len(patterns_to_save)}개 패턴이 화이트리스트에 저장되었습니다.")
            else:
                messagebox.showinfo("저장 정보", "저장할 사용자 수정 사항이 없습니다.")
                
        except Exception as e:
            messagebox.showerror("저장 오류", f"화이트리스트 저장 중 오류가 발생했습니다:\n{e}")
    
    def _show_help(self, event=None):
        """도움말 표시"""
        help_text = """
워크시트 분류 확인 - 도움말

🎯 목적:
전기 조서에서 백데이터(롤포워딩 대상) 워크시트를 식별합니다.

🤖 AI 자동 감지:
- 워크시트 이름 패턴 분석
- 데이터 구조 분석  
- 메타데이터 분석
- 3단계 종합 판정으로 95%+ 정확도 목표

📊 신뢰도 색상:
🟢 초록 (90%+): 높은 신뢰도 - AI 추천을 그대로 사용 권장
🟡 노랑 (50~90%): 중간 신뢰도 - 검토 후 결정 권장
🔴 빨강 (50% 미만): 낮은 신뢰도 - 수동 판단 필요

⌨️ 키보드 단축키:
- Tab: 다음 항목으로 이동
- Space: 체크박스 토글
- Enter: 미리보기 또는 확인
- Esc: 취소
- F1: 이 도움말 표시

💾 설정 저장:
사용자가 수정한 선택사항을 화이트리스트에 저장하여
다음번에 자동으로 적용할 수 있습니다.
"""
        
        help_window = tk.Toplevel(self.root)
        help_window.title("도움말")
        help_window.geometry("600x500")
        help_window.transient(self.root)
        
        text_area = scrolledtext.ScrolledText(help_window, 
            wrap=tk.WORD, font=("맑은 고딕", 10), padx=10, pady=10)
        text_area.pack(fill="both", expand=True)
        text_area.insert("1.0", help_text)
        text_area.config(state="disabled")
        
        close_btn = ttk.Button(help_window, text="닫기", 
            command=help_window.destroy)
        close_btn.pack(pady=10)
    
    def _on_key_press(self, event):
        """키 입력 처리"""
        if event.keysym == "F1":
            self._show_help()
            return "break"
        
        # Space로 현재 포커스된 체크박스 토글
        if event.keysym == "space":
            focused_widget = self.root.focus_get()
            # 체크박스에 포커스가 있으면 토글
            for widget_group in self.worksheet_widgets:
                if widget_group['checkbox'] == focused_widget:
                    current_value = widget_group['var'].get()
                    widget_group['var'].set(not current_value)
                    self._on_worksheet_toggled(widget_group['choice'], widget_group['var'])
                    return "break"
    
    def _on_tab_navigation(self, event):
        """Tab 네비게이션 처리"""
        # 기본 Tab 동작 허용
        return None
    
    def _on_shift_tab_navigation(self, event):
        """Shift+Tab 네비게이션 처리"""
        # 기본 Shift+Tab 동작 허용  
        return None
    
    def _on_confirm(self):
        """확인 버튼 클릭"""
        selected_count = sum(1 for choice in self.worksheet_choices if choice.is_selected)
        
        if selected_count == 0:
            result = messagebox.askyesno("확인", 
                "본 조서가 선택되지 않았습니다.\n"
                "모든 워크시트를 백데이터로 처리하시겠습니까?")
            if not result:
                return
        
        self.dialog_result = True
        self.root.destroy()
    
    def _on_cancel(self):
        """취소 버튼 클릭"""
        result = messagebox.askyesno("확인", "작업을 취소하시겠습니까?")
        if result:
            self.dialog_result = False
            self.root.destroy()
    
    def _process_results(self) -> Tuple[List[str], List[str]]:
        """결과 처리 및 반환"""
        main_worksheets = []
        back_data_worksheets = []
        
        for choice in self.worksheet_choices:
            if choice.is_selected:
                main_worksheets.append(choice.worksheet_name)
            else:
                back_data_worksheets.append(choice.worksheet_name)
        
        return main_worksheets, back_data_worksheets


def show_worksheet_confirmation(excel_path: Path, parent=None) -> Tuple[List[str], List[str]]:
    """
    워크시트 확인 대화상자 표시 (편의 함수)
    
    Args:
        excel_path: Excel 파일 경로
        parent: 부모 창 (선택사항)
        
    Returns:
        Tuple[List[str], List[str]]: (본조서 워크시트들, 백데이터 워크시트들)
    """
    dialog = WorksheetConfirmationDialog(excel_path, parent)
    return dialog.show_confirmation_dialog()


if __name__ == "__main__":
    # 테스트 코드
    import sys
    if len(sys.argv) > 1:
        test_file = Path(sys.argv[1])
        if test_file.exists():
            main_ws, back_ws = show_worksheet_confirmation(test_file)
            print(f"본 조서: {main_ws}")
            print(f"백데이터: {back_ws}")
        else:
            print(f"파일을 찾을 수 없습니다: {test_file}")
    else:
        print("사용법: python user_confirmation.py <excel_file_path>")